"""Analyze delivery settlement files for detailed fee breakdown."""
import openpyxl
import pandas as pd

# ═══════════════════════════════════════════
# 1. 쿠팡이츠
# ═══════════════════════════════════════════
print("=" * 60)
print("1. 쿠팡이츠 2월")
print("=" * 60)
wb = openpyxl.load_workbook(r"C:\WORK\SodamFN\2026소득분석\매출\2월\coupang_eats_2026-02.xlsx")
ws = wb.active

# Build column labels from merged header rows 1-2
col_labels = {}
for j in range(ws.max_column):
    r1 = str(ws.cell(1, j+1).value or "").strip()
    r2 = str(ws.cell(2, j+1).value or "").strip()
    if r1 or r2:
        col_labels[j] = f"{r1} > {r2}" if r1 and r2 else (r1 or r2)

# Sum all numeric columns
col_sums = {j: 0 for j in col_labels}
row_count = 0
for i in range(3, ws.max_row + 1):
    dt = ws.cell(i, 1).value
    if dt is None:
        continue
    row_count += 1
    for j in col_labels:
        v = ws.cell(i, j+1).value
        if v is not None:
            try:
                col_sums[j] += float(v)
            except:
                pass

print(f"총 {row_count}건")
print("-" * 50)
# Show fee breakdown (skip order info cols 0-8)
for j in sorted(col_labels.keys()):
    if j < 9:
        continue
    v = int(col_sums.get(j, 0))
    if v != 0:
        print(f"  {col_labels[j]}: {v:>15,}원")

# Calculate net
sales = int(col_sums.get(10, 0))  # 주문금액
settlement_col = None
for j in col_labels:
    if "정산금액" in col_labels[j] and "산정후" in col_labels[j]:
        settlement_col = j
# Try col 39 (정산금액/산정후)
settlement = int(col_sums.get(39, 0))
print(f"\n  주문금액(매출): {sales:>15,}원")
print(f"  정산금액: {settlement:>15,}원")
print(f"  총 수수료: {(sales - settlement):>15,}원")

# ═══════════════════════════════════════════
# 2. 배달의민족
# ═══════════════════════════════════════════
print("\n" + "=" * 60)
print("2. 배달의민족 2월")
print("=" * 60)
wb = openpyxl.load_workbook(r"C:\WORK\SodamFN\2026소득분석\매출\2월\배달의민족_2026년 2월 정산명세서.xlsx")

# Check 상세 sheet
ws = wb["상세"]
# Multi-row headers: rows 3-6
all_headers = {}
for j in range(26):
    parts = []
    for r in range(3, 7):
        v = ws.cell(r, j+1).value
        if v is not None:
            parts.append(str(v).strip())
    if parts:
        all_headers[j] = " > ".join(parts)

# Sum data rows
col_sums = {j: 0 for j in all_headers}
for i in range(7, ws.max_row + 1):
    for j in all_headers:
        v = ws.cell(i, j+1).value
        if v is not None:
            try:
                col_sums[j] += float(v)
            except:
                pass

print("컬럼별 합계:")
for j in sorted(all_headers.keys()):
    v = int(col_sums.get(j, 0))
    if v != 0:
        print(f"  {all_headers[j]}: {v:>15,}원")

# ═══════════════════════════════════════════
# 3. 요기요
# ═══════════════════════════════════════════
print("\n" + "=" * 60)
print("3. 요기요 2월")
print("=" * 60)
df = pd.read_excel(r"C:\WORK\SodamFN\2026소득분석\매출\2월\요기요_2026년_02월_정산내역_639-12-01514.xlsx")

# Show columns
print("컬럼 목록:")
for i, c in enumerate(df.columns):
    print(f"  Col {i}: {c}")

# Sum numeric columns
print("\n컬럼별 합계:")
for c in df.columns:
    if df[c].dtype in ["int64", "float64"]:
        v = int(df[c].sum())
        if v != 0:
            print(f"  {c}: {v:>15,}원")

# ═══════════════════════════════════════════
# 4. 땡겨요
# ═══════════════════════════════════════════
print("\n" + "=" * 60)
print("4. 땡겨요 2월")
print("=" * 60)
df = pd.read_excel(r"C:\WORK\SodamFN\2026소득분석\매출\2월\땡겨요 정산내역(2월건별).xls")

# Show first few rows to understand structure
print("컬럼 목록:")
for i, c in enumerate(df.columns):
    print(f"  Col {i}: {c}")

# Find header row
header_row = None
for i in range(min(50, len(df))):
    row_text = " ".join([str(v) for v in df.iloc[i].tolist() if pd.notna(v)])
    if "주문금액" in row_text and ("정산금액" in row_text or "정산" in row_text):
        header_row = i
        print(f"\nHeader found at row {i}:")
        for j, v in enumerate(df.iloc[i].tolist()):
            if pd.notna(v):
                print(f"  Col {j}: {v}")
        break

if header_row is not None:
    # Sum data rows
    col_sums = {}
    for i in range(header_row + 1, len(df)):
        first = str(df.iloc[i, 0]).strip() if pd.notna(df.iloc[i, 0]) else ""
        if "합 계" in first or "합계" in first:
            break
        for j in range(len(df.columns)):
            v = df.iloc[i, j]
            if pd.notna(v):
                try:
                    v = float(v)
                    header = str(df.iloc[header_row, j]) if pd.notna(df.iloc[header_row, j]) else f"Col{j}"
                    col_sums[header] = col_sums.get(header, 0) + v
                except:
                    pass
    
    print("\n컬럼별 합계:")
    for k, v in col_sums.items():
        iv = int(v)
        if iv != 0:
            print(f"  {k}: {iv:>15,}원")
