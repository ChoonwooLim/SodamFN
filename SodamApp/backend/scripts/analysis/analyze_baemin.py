"""
Analyze 배민 files with openpyxl instead of xlrd
"""
import openpyxl
import os

BASE_DIR = r"C:\WORK\SodamFN\2026소득분석\매출\배달앱매출\배민"

files = [
    "[배달의민족] HONG JI YEON 파트너님 2026년 1월 정산명세서.xlsx",
    "[배달의민족] HONG JI YEON 파트너님 2026년 2월 정산명세서.xlsx",
]

for fname in files:
    filepath = os.path.join(BASE_DIR, fname)
    print(f"\n{'='*60}")
    print(f"  FILE: {fname}")
    print(f"{'='*60}")
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n  Sheet: '{sheet_name}' — Rows: {ws.max_row}, Cols: {ws.max_column}")
            
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                vals = [str(v) if v is not None else '' for v in row]
                print(f"    Row {row_idx}: {vals}")
                if row_idx >= 60:
                    print("    ... (truncated)")
                    break
    except Exception as e:
        print(f"  ERROR: {e}")
        # Try with xlrd
        try:
            import xlrd
            wb2 = xlrd.open_workbook(filepath)
            for sheet_name in wb2.sheet_names():
                ws2 = wb2.sheet_by_name(sheet_name)
                print(f"\n  [xlrd] Sheet: '{sheet_name}' — Rows: {ws2.nrows}, Cols: {ws2.ncols}")
                for r in range(min(60, ws2.nrows)):
                    vals = [str(ws2.cell(r, c).value) for c in range(ws2.ncols)]
                    print(f"    Row {r}: {vals}")
        except Exception as e2:
            print(f"  xlrd ERROR: {e2}")
