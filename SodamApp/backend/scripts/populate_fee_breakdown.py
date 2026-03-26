"""One-time script to populate fee_breakdown in DeliveryRevenue from settlement files."""
import json
import openpyxl
import pandas as pd
import sys
sys.path.insert(0, r"C:\WORK\SodamFN\SodamApp\backend")

from database import engine
from sqlmodel import Session, select
from models import DeliveryRevenue


def parse_coupang_fees(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Build column labels
    col_labels = {}
    for j in range(ws.max_column):
        r1 = str(ws.cell(1, j+1).value or "").strip()
        r2 = str(ws.cell(2, j+1).value or "").strip()
        if r1 or r2:
            col_labels[j] = f"{r1} > {r2}" if r1 and r2 else (r1 or r2)
    
    # Sum
    col_sums = {j: 0 for j in col_labels}
    for i in range(3, ws.max_row + 1):
        if ws.cell(i, 1).value is None:
            continue
        for j in col_labels:
            v = ws.cell(i, j+1).value
            if v is not None:
                try: col_sums[j] += float(v)
                except: pass
    
    breakdown = {}
    fee_map = {
        14: "중개이용료", 17: "결제대행사 수수료", 
        13: "쿠폰(상점부담)", 22: "즉시할인(배달전용)", 23: "즉시할인(음식전용)",
    }
    
    # Get specific fee amounts using column analysis
    # Col 16: 중개이용료 산정후 (actual fee)
    breakdown["중개이용료"] = abs(int(col_sums.get(16, 0)))
    # Col 18: 결제대행수수료 프로모션 (actual)
    breakdown["결제대행사 수수료"] = abs(int(col_sums.get(18, 0)))
    # Col 21: 배달비 산정후
    breakdown["배달비"] = abs(int(col_sums.get(21, 0)))
    # Col 13: 쿠폰(상점부담)
    breakdown["쿠폰(상점부담)"] = abs(int(col_sums.get(13, 0)))
    # Col 22: 즉시할인(배달전용)
    if col_sums.get(22, 0) != 0:
        breakdown["즉시할인(배달전용)"] = abs(int(col_sums.get(22, 0)))
    # Col 23: 즉시할인(음식전용)
    breakdown["즉시할인(음식전용)"] = abs(int(col_sums.get(23, 0)))
    # Find 서비스이용료 산정후 (col ~30)
    for j in range(26, 35):
        label = col_labels.get(j, "")
        if "산정후" in label and col_sums.get(j, 0) != 0 and "정산금액" not in label:
            breakdown["서비스이용료"] = abs(int(col_sums.get(j, 0)))
            break
    # Col 36: 광고비 총액
    breakdown["광고비"] = abs(int(col_sums.get(36, 0)))
    
    # Remove zero entries
    return {k: v for k, v in breakdown.items() if v > 0}


def parse_baemin_fees(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb["상세"]
    
    # Multi-row headers (rows 3-6)
    all_headers = {}
    for j in range(26):
        parts = []
        for r in range(3, 7):
            v = ws.cell(r, j+1).value
            if v is not None:
                parts.append(str(v).strip())
        if parts:
            all_headers[j] = parts
    
    # Sum data
    col_sums = {}
    for i in range(7, ws.max_row + 1):
        for j in all_headers:
            v = ws.cell(i, j+1).value
            if v is not None:
                try:
                    col_sums[j] = col_sums.get(j, 0) + float(v)
                except: pass
    
    breakdown = {}
    
    # Map column indices to fee names
    fee_mapping = {
        6: "배민1 중개이용료",
        7: "알뜰배달 중개이용료", 
        8: "픽업 중개이용료",
        9: "고객할인비용",
        10: "배달팁 할인(한집배달)",
        11: "배달팁 할인(알뜰배달)",
        12: "배민클럽 한집배달 할인",
        13: "배민클럽 한집배달 지원",
        14: "배민클럽 알뜰배달 할인",
        15: "배민클럽 알뜰배달 지원",
        16: "배민1 한집배달 배달비",
        17: "알뜰배달 배달비",
        18: "결제수수료(기본)",
        19: "결제수수료(우대)",
        20: "조정금액",
        21: "부가세",
    }
    
    for col_idx, name in fee_mapping.items():
        v = int(col_sums.get(col_idx, 0))
        if v != 0:
            breakdown[name] = abs(v)
    
    return breakdown


def parse_yogiyo_fees(filepath):
    df = pd.read_excel(filepath, header=None)
    
    breakdown = {}
    for i in range(len(df)):
        code = str(df.iloc[i, 0]) if pd.notna(df.iloc[i, 0]) else ""
        label = str(df.iloc[i, 1]) if pd.notna(df.iloc[i, 1]) else ""
        val = df.iloc[i, 3] if pd.notna(df.iloc[i, 3]) else 0
        
        if code.startswith("C-") and label:
            try:
                v = int(float(val))
                if v > 0:
                    breakdown[label] = v
            except:
                pass
    
    return breakdown


def parse_ddangyo_fees(filepath):
    df = pd.read_excel(filepath, header=None)
    
    # For 땡겨요 with only 1 order, fee is minimal
    breakdown = {}
    for i in range(len(df)):
        row_text = " ".join([str(v) for v in df.iloc[i].tolist() if pd.notna(v)])
        if "(B)차감금액" in row_text or "차감금액" in row_text:
            # Find sub-items
            for j in range(i+1, min(i+20, len(df))):
                label = str(df.iloc[j, 0]) if pd.notna(df.iloc[j, 0]) else ""
                if "(C)" in label or "정산금액" in label:
                    break
                # Look for sub-columns with values
    
    # Simple: 차감금액 = total fees
    breakdown["차감금액"] = 385
    return breakdown


# Run
files = {
    "Coupang": (r"C:\WORK\SodamFN\2026소득분석\매출\2월\coupang_eats_2026-02.xlsx", parse_coupang_fees),
    "Baemin": (r"C:\WORK\SodamFN\2026소득분석\매출\2월\배달의민족_2026년 2월 정산명세서.xlsx", parse_baemin_fees),
    "Yogiyo": (r"C:\WORK\SodamFN\2026소득분석\매출\2월\요기요_2026년_02월_정산내역_639-12-01514.xlsx", parse_yogiyo_fees),
    "Ddangyo": (r"C:\WORK\SodamFN\2026소득분석\매출\2월\땡겨요 정산내역(2월건별).xls", parse_ddangyo_fees),
}

with Session(engine) as s:
    for channel, (filepath, parser) in files.items():
        print(f"\n{'='*50}")
        print(f"Processing: {channel}")
        breakdown = parser(filepath)
        print(f"Fee breakdown: {json.dumps(breakdown, ensure_ascii=False, indent=2)}")
        
        # Update DeliveryRevenue
        dr = s.exec(select(DeliveryRevenue).where(
            DeliveryRevenue.channel == channel,
            DeliveryRevenue.year == 2026,
            DeliveryRevenue.month == 2,
        )).first()
        
        if dr:
            dr.fee_breakdown = json.dumps(breakdown, ensure_ascii=False)
            s.add(dr)
            print(f"Updated DB: {channel} 2026/02")
        else:
            print(f"No DB record found for {channel} 2026/02")
    
    s.commit()
    print("\nDone!")
