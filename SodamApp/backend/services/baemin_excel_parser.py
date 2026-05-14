"""배민(self.baemin.com) 정산명세서 엑셀 파서 — Phase 2a 수동 import.

self-api.baemin.com 자동수집은 동적 signature 인증 차단으로 무용. 사장님이 매월
배민 사이트에서 정산명세서 xlsx 를 수동 다운로드 → 셈하나 UI 에 업로드.

다운로드 출처:
  배민 사장님사이트 → 정산 → 정산명세서 (월 단위)

비밀번호:
  매월 동일 패스워드(예: 사장님 자체 설정). msoffcrypto 로 복호화 시도, 실패 시 평문 시도.
  2026-02 부터는 평문으로 다운로드되는 케이스도 발생 → fallback 필수.

시트 구조:
  [요약]
    R1: '2026년 N월 정산명세서'
    R2: (빈 행)
    R3: '정산 요약'
    R4: 설명 문구
    R5: ['(A) 주문중개', '(B) 배달', '(C) 그외', '(D) 기타',
         '(E) 부가세', '(F) 우리가게클릭', '(G) 배민오더', '(H) 입금금액']
    R6: 합계 값 (입금완료건만)

  [상세]
    R1: '정산 상세내역'
    R2: '거래기간: YYYY-MM-DD ~ YYYY-MM-DD'
    R3: 그룹 헤더 — '입금일', '정산대상기간', '입금 금액', '이용서비스', '주문유형/기타',
                 '(A) 주문중개', '(B) 배달', '(C) 그외', '(D) 기타', '(E) 부가세',
                 '(F) 우리가게클릭', '(G) 배민오더', '(H) 입금금액', '상태'
    R4: 중분류 — '주문금액', '중개이용료', '고객할인비용', '배달팁 할인비용',
                '배민클럽 할인비용', '배달비', '결제정산수수료', '조정금액' 등
    R5: 리프 컬럼명 (실제 컬럼명) — '바로결제주문금액', '배민1중개이용료',
                                  '알뜰배달 중개이용료', '픽업중개이용료',
                                  '주문금액 즉시할인', '한집배달 배달팁 즉시할인',
                                  '기본수수료(정률)', '우대수수료' 등
    R6~: 데이터 (각 row = 입금 단위)

매핑 전략:
  - R5 leaf 가 있으면 R5 leaf 텍스트 매칭
  - R5 가 None 인 컬럼 (예: 입금일/상태 — single-level): R3 그룹 텍스트 매칭
  - R5 가 None 이고 R4 만 있는 컬럼 (예: (D) 기타: '조정금액'): R4 텍스트 매칭
  - 동명이의(2개 '부가세') 충돌 회피: R3 그룹 제약 — (E) vs (F) 우리가게클릭

컬럼 수 차이 (월별):
  1월: 27 cols — '부분환불금액' 있음, 우리가게클릭 단일컬럼
  2월: 26 cols — '부분환불금액' 없음, 조정금액 있음
  3월: 28 cols — '부분환불금액' 있음, 우리가게클릭 → 요금+부가세 2컬럼
"""
from __future__ import annotations

import datetime
import io
import json
import logging
from dataclasses import dataclass, field
from typing import Optional

import openpyxl

log = logging.getLogger("baemin.excel")


class BaeminExcelError(Exception):
    """배민 정산명세서 엑셀 파싱 실패."""


# ──────────────────────────────────────────────────────────────────────────
# Data classes
# ──────────────────────────────────────────────────────────────────────────


@dataclass
class ParsedBaeminRow:
    """[상세] 시트 1행 — 입금 단위."""

    # 기본 정보
    deposit_date: Optional[datetime.date] = None
    settlement_period: Optional[str] = None
    deposit_amount: int = 0
    service_type: Optional[str] = None
    order_type: Optional[str] = None

    # (A) 주문중개
    order_amount: int = 0
    refund_amount: int = 0
    brokerage_baemin1: int = 0
    brokerage_smart: int = 0
    brokerage_pickup: int = 0
    customer_discount: int = 0

    # (B) 배달
    tip_discount_single: int = 0
    tip_discount_smart: int = 0
    club_single_discount: int = 0
    club_single_subsidy: int = 0
    club_smart_discount: int = 0
    club_smart_subsidy: int = 0
    delivery_fee_single: int = 0
    delivery_fee_smart: int = 0

    # (C) 그외
    payment_fee_base: int = 0
    payment_fee_preferred: int = 0

    # (D) 기타 / 조정
    etc_amount: int = 0
    adjustment_amount: int = 0

    # (E)~(H)
    vat: int = 0
    ad_amount: int = 0
    ad_vat: int = 0
    baemin_order_amount: int = 0
    deposit_final: int = 0
    status: Optional[str] = None

    # 메타
    raw_row: tuple = field(default_factory=tuple)

    @property
    def is_completed(self) -> bool:
        return (self.status or "").strip() == "입금완료"


@dataclass
class ParsedBaeminMonth:
    """배민 정산명세서 한 달 — [요약] + [상세] 통합."""
    summary: dict = field(default_factory=dict)
    detail_rows: list[ParsedBaeminRow] = field(default_factory=list)
    file_name: Optional[str] = None
    detail_columns: int = 0          # 상세 시트 max_column (디버그)
    column_index_map: dict = field(default_factory=dict)
    period_text: Optional[str] = None   # '거래기간: 2026-02-25 ~ 2026-03-26'


# ──────────────────────────────────────────────────────────────────────────
# 변환 헬퍼
# ──────────────────────────────────────────────────────────────────────────


_DATE_FORMATS = ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d")


def _to_date(value) -> Optional[datetime.date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    s = str(value).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_str(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _to_int(value) -> int:
    """엑셀 셀 값 → int (음수/0/-0/None/문자열 다 안전)."""
    if value is None or value == "":
        return 0
    if isinstance(value, bool):
        return 0
    if isinstance(value, (int, float)):
        try:
            return int(round(float(value)))
        except (OverflowError, ValueError):
            return 0
    s = str(value).strip().replace(",", "").replace(" ", "")
    if not s:
        return 0
    try:
        return int(round(float(s)))
    except (ValueError, OverflowError):
        return 0


# ──────────────────────────────────────────────────────────────────────────
# 복호화
# ──────────────────────────────────────────────────────────────────────────


def decrypt_xlsx(file_bytes: bytes, password: Optional[str] = None) -> io.BytesIO:
    """msoffcrypto 로 복호화 시도. 실패하면 평문으로 간주하고 원본 buf 반환.

    배민 정산명세서는 보통 비밀번호 보호되지만 2026-02 부터는 평문 케이스도 발생.
    """
    if not file_bytes:
        raise BaeminExcelError("빈 파일")

    # 우선 복호화 시도
    if password:
        try:
            import msoffcrypto  # type: ignore
            in_buf = io.BytesIO(file_bytes)
            ofile = msoffcrypto.OfficeFile(in_buf)
            ofile.load_key(password=password)
            out = io.BytesIO()
            ofile.decrypt(out)
            out.seek(0)
            return out
        except Exception as e:  # noqa: BLE001
            # 평문이거나 비번 틀림 — 평문 시도로 fallback
            log.info("msoffcrypto decrypt failed (fallback to plaintext): %s", e)

    # 평문 시도
    bio = io.BytesIO(file_bytes)
    # 평문 검증 — PK 시그너처
    if file_bytes[:4] != b"PK\x03\x04":
        raise BaeminExcelError(
            "암호화된 파일이지만 비밀번호로 복호화 실패. "
            "비밀번호가 올바른지 확인해주세요."
        )
    return bio


# ──────────────────────────────────────────────────────────────────────────
# 컬럼 매핑
# ──────────────────────────────────────────────────────────────────────────


# 필드명 → 후보 헤더 텍스트 list. 첫 매칭 우선.
# group_constraint: R3 그룹 (예 '(F) 우리가게클릭') 으로 제약. None 이면 그룹 무시.
# label_row: 'r5_leaf' = R5 leaf 우선, 'r4_mid' = R4 중분류, 'r3_group' = R3 그룹 단일 셀
_FIELD_SPECS: dict[str, list[dict]] = {
    # ─ 좌측 single-level (R3 만 채워짐, R4·R5 = None)
    "deposit_date": [
        {"row": "r3", "labels": ["입금일"]},
    ],
    "settlement_period": [
        {"row": "r3", "labels": ["정산대상기간"]},
    ],
    "deposit_amount": [
        {"row": "r3", "labels": ["입금 금액"]},
    ],
    "service_type": [
        {"row": "r3", "labels": ["이용서비스"]},
    ],
    "order_type": [
        {"row": "r3", "labels": ["주문유형/기타"]},
    ],

    # ─ (A) 주문중개 leaf
    "order_amount": [
        {"row": "r5", "labels": ["바로결제주문금액"], "group": "(A) 주문중개"},
    ],
    "refund_amount": [
        {"row": "r4", "labels": ["부분환불금액"], "group": "(A) 주문중개"},
    ],
    "brokerage_baemin1": [
        {"row": "r5", "labels": ["배민1중개이용료"], "group": "(A) 주문중개"},
    ],
    "brokerage_smart": [
        {"row": "r5", "labels": ["알뜰배달 중개이용료"], "group": "(A) 주문중개"},
    ],
    "brokerage_pickup": [
        {"row": "r5", "labels": ["픽업중개이용료"], "group": "(A) 주문중개"},
    ],
    "customer_discount": [
        {"row": "r5", "labels": ["주문금액 즉시할인"], "group": "(A) 주문중개"},
    ],

    # ─ (B) 배달 leaf
    "tip_discount_single": [
        {"row": "r5", "labels": ["한집배달 배달팁 즉시할인"], "group": "(B) 배달"},
    ],
    "tip_discount_smart": [
        {"row": "r5", "labels": ["알뜰배달 배달팁 즉시할인"], "group": "(B) 배달"},
    ],
    "club_single_discount": [
        {"row": "r5", "labels": ["배민클럽(한집배달) 배달팁 할인"], "group": "(B) 배달"},
    ],
    "club_single_subsidy": [
        {"row": "r5", "labels": ["배민클럽(한집배달) 배달팁 할인 지원"], "group": "(B) 배달"},
    ],
    "club_smart_discount": [
        {"row": "r5", "labels": ["배민클럽(알뜰배달) 배달팁 할인"], "group": "(B) 배달"},
    ],
    "club_smart_subsidy": [
        {"row": "r5", "labels": ["배민클럽(알뜰배달) 배달팁 할인 지원"], "group": "(B) 배달"},
    ],
    "delivery_fee_single": [
        {"row": "r5", "labels": ["배민1 한집배달 배달비"], "group": "(B) 배달"},
    ],
    "delivery_fee_smart": [
        {"row": "r5", "labels": ["알뜰배달 배달비"], "group": "(B) 배달"},
    ],

    # ─ (C) 그외 (결제정산수수료)
    "payment_fee_base": [
        {"row": "r5", "labels": ["기본수수료(정률)"], "group": "(C) 그외"},
    ],
    "payment_fee_preferred": [
        {"row": "r5", "labels": ["우대수수료"], "group": "(C) 그외"},
    ],

    # ─ (D) 기타 — R5 가 None, R3 = '(D) 기타' (단일 셀 컬럼)
    "etc_amount": [
        # 1·3월: '(D) 기타' R3 single, R4·R5 모두 None
        # 2월: R3 = '(D) 기타', R4 = '조정금액', R5 = '보정금액' → adjustment_amount 로 분리
        # 여기선 R4/R5 가 모두 비어있는 R3 = '(D) 기타' 컬럼만 매칭
        {"row": "r3_blank_below", "labels": ["(D) 기타"]},
    ],
    "adjustment_amount": [
        # 2월: R4 = '조정금액', 3월 '조정금액' R4, '보정금액' R5
        {"row": "r5", "labels": ["보정금액"], "group": "(D) 기타"},
        {"row": "r4", "labels": ["조정금액"], "group": "(D) 기타"},
    ],

    # ─ (E) 부가세 — single column
    "vat": [
        {"row": "r3_blank_below", "labels": ["(E) 부가세"]},
    ],

    # ─ (F) 우리가게클릭
    # 1·2월: R3 = '(F) 우리가게클릭', R4·R5 모두 None → 단일 컬럼
    # 3월: R3 = '(F) 우리가게클릭' (병합 2칸), R4 = ['우리가게클릭 이용요금', '부가세'], R5 = None
    "ad_amount": [
        {"row": "r4", "labels": ["우리가게클릭 이용요금"], "group": "(F) 우리가게클릭"},
        {"row": "r3_blank_below", "labels": ["(F) 우리가게클릭"]},
    ],
    "ad_vat": [
        # 3월~ 만 존재. R4 = '부가세', R3 = '(F) 우리가게클릭' 그룹 제약 필수 (E 와 충돌)
        {"row": "r4", "labels": ["부가세"], "group": "(F) 우리가게클릭"},
    ],

    # ─ (G) 배민오더
    "baemin_order_amount": [
        {"row": "r3_blank_below", "labels": ["(G) 배민오더"]},
    ],

    # ─ (H) 입금금액
    "deposit_final": [
        {"row": "r3_blank_below", "labels": ["(H) 입금금액"]},
    ],

    # ─ 상태
    "status": [
        {"row": "r3", "labels": ["상태"]},
    ],
}


def _cell_str(v):
    if v is None or v == "":
        return None
    s = str(v).strip()
    return s or None


def _ffill(arr: list[Optional[str]]) -> list[Optional[str]]:
    out = []
    last = None
    for v in arr:
        if v is not None:
            last = v
        out.append(last)
    return out


def _build_column_index_map(ws) -> tuple[dict[str, int], list[dict]]:
    """[상세] 시트 R3(group) + R4(mid) + R5(leaf) → 필드명 → 컬럼 인덱스 매핑.

    Returns:
        (idx_map, debug_paths) — idx_map: 필드 → 컬럼 인덱스
                                debug_paths: 각 컬럼의 (r3, r4, r5) 디버그용
    """
    max_col = ws.max_column
    r3 = [_cell_str(ws.cell(row=3, column=c).value) for c in range(1, max_col + 1)]
    r4 = [_cell_str(ws.cell(row=4, column=c).value) for c in range(1, max_col + 1)]
    r5 = [_cell_str(ws.cell(row=5, column=c).value) for c in range(1, max_col + 1)]

    # R3, R4 는 병합셀일 가능성이 높음 (그룹 헤더). forward-fill 적용.
    r3_ff = _ffill(r3)
    r4_ff = _ffill(r4)
    # R5 는 leaf 라 ffill 안 함 (None 의미 있음)

    # 디버그 path
    debug_paths = [
        {"col": i, "r3": r3_ff[i], "r4": r4_ff[i], "r5": r5[i],
         "r3_raw": r3[i], "r4_raw": r4[i]}
        for i in range(max_col)
    ]

    idx_map: dict[str, int] = {}
    for fld, specs in _FIELD_SPECS.items():
        for spec in specs:
            row_kind = spec["row"]
            labels = spec["labels"]
            group = spec.get("group")
            found_idx: Optional[int] = None

            for i in range(max_col):
                # 그룹 제약 (R3 ffill 기준)
                if group is not None and r3_ff[i] != group:
                    continue

                if row_kind == "r3":
                    # R3 가 정확히 label 인 컬럼 (single-level — r3 raw cell, ffill 아님)
                    if r3[i] in labels:
                        found_idx = i
                        break
                elif row_kind == "r3_blank_below":
                    # R3 = label, R4/R5 모두 None — 단일컬럼 group (e.g. (E)부가세, (H)입금금액).
                    # R3 raw 가 label 이거나, R3 ffill 이 label 이면서 r4/r5 둘 다 None
                    if (r3[i] in labels) and r4[i] is None and r5[i] is None:
                        found_idx = i
                        break
                    if r3_ff[i] in labels and r4[i] is None and r5[i] is None:
                        found_idx = i
                        break
                elif row_kind == "r4":
                    # R4 = label (그룹 제약 함께)
                    # raw r4 셀이 label 일 때만 매칭 (ffill 은 다른 컬럼에 흘러올 수 있음)
                    if r4[i] in labels:
                        found_idx = i
                        break
                elif row_kind == "r5":
                    if r5[i] in labels:
                        found_idx = i
                        break

            if found_idx is not None:
                idx_map[fld] = found_idx
                break

    return idx_map, debug_paths


# ──────────────────────────────────────────────────────────────────────────
# 파서 본체
# ──────────────────────────────────────────────────────────────────────────


def parse_summary(xlsx_buf: io.BytesIO) -> dict:
    """[요약] 시트의 R6 (합계) 값 8개 → dict.

    Returns:
        {
          'order_brokerage_total': int,
          'delivery_total': int,
          'etc_total': int,
          'misc_total': int,
          'vat_total': int,
          'ad_total': int,
          'baemin_order_total': int,
          'deposit_total': int,
        }
    """
    xlsx_buf.seek(0)
    wb = openpyxl.load_workbook(xlsx_buf, data_only=True)
    if "요약" not in wb.sheetnames:
        raise BaeminExcelError(f"'요약' 시트를 찾을 수 없습니다. sheets={wb.sheetnames}")
    ws = wb["요약"]
    # R5 = 컬럼 라벨, R6 = 값
    labels = [_cell_str(ws.cell(row=5, column=c).value) for c in range(1, 9)]
    values = [_to_int(ws.cell(row=6, column=c).value) for c in range(1, 9)]

    expected = ["(A) 주문중개", "(B) 배달", "(C) 그외", "(D) 기타",
                "(E) 부가세", "(F) 우리가게클릭", "(G) 배민오더", "(H) 입금금액"]
    # 위치 기반이 더 안정 (라벨이 한칸 변동될 가능성 낮음)
    if labels[:8] != expected[:8]:
        log.warning(
            "요약 R5 라벨이 예상과 다름. actual=%s expected=%s",
            labels[:8], expected[:8],
        )

    return {
        "order_brokerage_total": values[0],
        "delivery_total": values[1],
        "etc_total": values[2],
        "misc_total": values[3],
        "vat_total": values[4],
        "ad_total": values[5],
        "baemin_order_total": values[6],
        "deposit_total": values[7],
    }


def parse_detail(xlsx_buf: io.BytesIO
                 ) -> tuple[list[ParsedBaeminRow], dict[str, int], int, Optional[str]]:
    """[상세] 시트 → rows + index map + col count + 거래기간 텍스트.

    Returns:
        (rows, idx_map, max_column, period_text)
    """
    xlsx_buf.seek(0)
    wb = openpyxl.load_workbook(xlsx_buf, data_only=True)
    if "상세" not in wb.sheetnames:
        raise BaeminExcelError(f"'상세' 시트를 찾을 수 없습니다. sheets={wb.sheetnames}")
    ws = wb["상세"]

    period_text = _cell_str(ws.cell(row=2, column=1).value)

    if ws.max_column < 20:
        raise BaeminExcelError(
            f"상세 컬럼 수 {ws.max_column} 너무 적음 — 포맷 깨짐 의심."
        )

    idx_map, debug_paths = _build_column_index_map(ws)

    # 필수 필드 검증
    required = ("deposit_date", "deposit_amount", "order_amount",
                "deposit_final", "status")
    missing = [f for f in required if f not in idx_map]
    if missing:
        dump = "\n".join(
            f"  col[{p['col']:2d}] R3='{p['r3_raw']}' R4='{p['r4_raw']}' R5='{p['r5']}'"
            for p in debug_paths
        )
        raise BaeminExcelError(
            f"필수 컬럼 매핑 실패 {missing} — max_col={ws.max_column}\n"
            f"감지된 헤더:\n{dump}"
        )

    log.info(
        "baemin excel column map max_col=%d mapped=%d/%d",
        ws.max_column, len(idx_map), len(_FIELD_SPECS),
    )

    rows: list[ParsedBaeminRow] = []
    for r_idx in range(6, ws.max_row + 1):
        row_vals = tuple(ws.cell(row=r_idx, column=c).value
                         for c in range(1, ws.max_column + 1))
        # 빈 행 (전체 None) 스킵
        if all(v is None or v == "" for v in row_vals):
            continue
        # 입금일 컬럼이 비어있으면 데이터 row 가 아님 (footer 등)
        deposit_idx = idx_map.get("deposit_date")
        if deposit_idx is None or deposit_idx >= len(row_vals):
            continue
        dd = _to_date(row_vals[deposit_idx])
        if dd is None:
            # 입금일 파싱 실패 — footer 등 데이터 아닌 행
            continue

        def gi(field_name: str) -> int:
            idx = idx_map.get(field_name)
            if idx is None or idx >= len(row_vals):
                return 0
            return _to_int(row_vals[idx])

        def gs(field_name: str) -> Optional[str]:
            idx = idx_map.get(field_name)
            if idx is None or idx >= len(row_vals):
                return None
            return _to_str(row_vals[idx])

        rows.append(ParsedBaeminRow(
            deposit_date=dd,
            settlement_period=gs("settlement_period"),
            deposit_amount=gi("deposit_amount"),
            service_type=gs("service_type"),
            order_type=gs("order_type"),
            order_amount=gi("order_amount"),
            refund_amount=gi("refund_amount"),
            brokerage_baemin1=gi("brokerage_baemin1"),
            brokerage_smart=gi("brokerage_smart"),
            brokerage_pickup=gi("brokerage_pickup"),
            customer_discount=gi("customer_discount"),
            tip_discount_single=gi("tip_discount_single"),
            tip_discount_smart=gi("tip_discount_smart"),
            club_single_discount=gi("club_single_discount"),
            club_single_subsidy=gi("club_single_subsidy"),
            club_smart_discount=gi("club_smart_discount"),
            club_smart_subsidy=gi("club_smart_subsidy"),
            delivery_fee_single=gi("delivery_fee_single"),
            delivery_fee_smart=gi("delivery_fee_smart"),
            payment_fee_base=gi("payment_fee_base"),
            payment_fee_preferred=gi("payment_fee_preferred"),
            etc_amount=gi("etc_amount"),
            adjustment_amount=gi("adjustment_amount"),
            vat=gi("vat"),
            ad_amount=gi("ad_amount"),
            ad_vat=gi("ad_vat"),
            baemin_order_amount=gi("baemin_order_amount"),
            deposit_final=gi("deposit_final"),
            status=gs("status"),
            raw_row=row_vals,
        ))

    return rows, idx_map, ws.max_column, period_text


def parse_xlsx(file_bytes: bytes,
               password: Optional[str] = None,
               file_name: Optional[str] = None) -> ParsedBaeminMonth:
    """진입점 — 파일 bytes → ParsedBaeminMonth.

    1) 비번 복호화 시도 (실패 시 평문)
    2) [요약] + [상세] 파싱
    3) ParsedBaeminMonth 반환
    """
    buf = decrypt_xlsx(file_bytes, password=password)
    summary = parse_summary(buf)
    rows, idx_map, max_col, period_text = parse_detail(buf)
    return ParsedBaeminMonth(
        summary=summary,
        detail_rows=rows,
        file_name=file_name,
        detail_columns=max_col,
        column_index_map=idx_map,
        period_text=period_text,
    )


# ──────────────────────────────────────────────────────────────────────────
# 집계 (DeliveryRevenue 갱신용)
# ──────────────────────────────────────────────────────────────────────────


@dataclass
class CompletedAggregate:
    """입금완료 row 만 합산 — DeliveryRevenue 갱신용."""
    row_count: int = 0
    total_order_amount: int = 0     # 바로결제주문금액 합 (총매출)
    total_refund: int = 0
    total_deposit: int = 0          # (H) 입금금액 합 (실수령)
    total_deduction: int = 0        # 총 차감 = order_amount - deposit_final


def aggregate_completed(detail_rows: list[ParsedBaeminRow]) -> CompletedAggregate:
    """status='입금완료' row 만 합산."""
    agg = CompletedAggregate()
    for r in detail_rows:
        if not r.is_completed:
            continue
        agg.row_count += 1
        agg.total_order_amount += r.order_amount
        agg.total_refund += r.refund_amount
        agg.total_deposit += r.deposit_final
    agg.total_deduction = agg.total_order_amount - agg.total_deposit
    return agg


def row_to_raw_json(row: ParsedBaeminRow) -> str:
    """ParsedBaeminRow.raw_row → JSON string. datetime/date 는 isoformat."""
    def _enc(v):
        if isinstance(v, (datetime.date, datetime.datetime)):
            return v.isoformat()
        return v
    try:
        return json.dumps([_enc(v) for v in row.raw_row], ensure_ascii=False)
    except Exception:  # noqa: BLE001
        return json.dumps([str(v) for v in row.raw_row], ensure_ascii=False)
