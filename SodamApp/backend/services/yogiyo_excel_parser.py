"""요기요 정산내역 엑셀(.xlsx) 파서.

구조 (실제 샘플 기준):
  [요약] 시트: 코드가 col0, 항목명 col1, 값이 그 행의 마지막 숫자셀.
    A  주문금액   (= 매출 gross)
    B  일회용컵보증금
    C  차감금액   (= 총비용, 수수료 전액)
    C-1 ~ C-14    (차감 항목별: 할인·중개·배달대행·외부결제·광고 등)
    D  정산금액   (= 정산)
    F  입금받으실 금액
  [상세 거래내역] 시트: r2 헤더, r3+ 주문 1건/행 (NO 컬럼).

주의: 요기요 정산내역도 지급일(정산) 기준이므로, 화면 매출은 결정 A(주문 기준
DailyExpense)를 쓰고 여기서는 **수수료율 산출용 총비용/정산/gross** 만 제공한다.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Optional

import openpyxl


class YogiyoExcelError(Exception):
    pass


@dataclass
class ParsedYogiyoMonth:
    gross: int = 0             # A 주문금액
    total_fees: int = 0        # C 차감금액 (총비용)
    settlement: int = 0        # D 정산금액
    order_count: int = 0       # 상세 거래내역 행 수
    fee_breakdown: dict = field(default_factory=dict)  # {항목명: 금액} (C-x 중 non-zero)

    @property
    def fee_rate(self) -> float:
        return round(self.total_fees / self.gross * 100, 1) if self.gross > 0 else 0.0


def _last_number(row) -> Optional[float]:
    """행에서 마지막 숫자 셀 값 반환 (없으면 None)."""
    val = None
    for c in row:
        if isinstance(c, (int, float)) and not isinstance(c, bool):
            val = c
    return val


def _to_int(v) -> int:
    try:
        return int(round(abs(float(v))))
    except (TypeError, ValueError):
        return 0


def parse_xlsx(file_bytes: bytes) -> ParsedYogiyoMonth:
    """요기요 정산내역 xlsx bytes → ParsedYogiyoMonth."""
    if not file_bytes or file_bytes[:4] != b"PK\x03\x04":
        raise YogiyoExcelError("xlsx(.xlsx) 파일이 아닙니다.")
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # 요약 시트 찾기 (이름에 '요약' 포함, 없으면 첫 시트)
    summary_ws = None
    for ws in wb.worksheets:
        if "요약" in ws.title:
            summary_ws = ws
            break
    summary_ws = summary_ws or wb.worksheets[0]

    codes = {}          # 코드 → 값
    names = {}          # 코드 → 항목명
    for row in summary_ws.iter_rows(values_only=True):
        if not row:
            continue
        code = ("" if row[0] is None else str(row[0])).strip()
        if not code:
            continue
        num = _last_number(row)
        if num is None:
            continue
        codes[code] = num
        names[code] = ("" if len(row) < 2 or row[1] is None else str(row[1])).strip()

    if "A" not in codes or "C" not in codes or "D" not in codes:
        raise YogiyoExcelError(
            f"요약 시트에서 A/C/D 코드를 찾지 못했습니다. 발견={sorted(codes)[:10]}"
        )

    result = ParsedYogiyoMonth(
        gross=_to_int(codes.get("A")),
        total_fees=_to_int(codes.get("C")),
        settlement=_to_int(codes.get("D")),
    )
    for code, val in codes.items():
        if code.startswith("C-") and _to_int(val) > 0:
            result.fee_breakdown[names.get(code) or code] = _to_int(val)

    # C-x 항목 중 환급(양수 표기) 등이 절댓값 처리로 과대 합산될 수 있음 —
    # 잔차를 기타조정으로 흡수해 breakdown 합계 == C 차감금액 을 항상 보장.
    if result.fee_breakdown:
        resid = result.total_fees - sum(result.fee_breakdown.values())
        if resid != 0:
            result.fee_breakdown["기타조정"] = resid

    # 상세 거래내역 시트에서 주문 건수
    for ws in wb.worksheets:
        if "상세" in ws.title:
            cnt = 0
            for row in ws.iter_rows(min_row=1, values_only=True):
                if row and isinstance(row[0], (int, float)) and not isinstance(row[0], bool):
                    cnt += 1
            result.order_count = cnt
            break

    return result
