"""쿠팡이츠 월별 매출내역서 Excel (43컬럼) 파서.

다운로드 출처:
  GET /api/v1/merchant/web/emails
      ?type=salesOrder&action=download&downloadRequestDate=YYYY-MM&storeId=...

응답: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
시트 이름: 'YYYY-MM_1'
구조:
  row 1: 그룹 헤더 (매출액 / 쿠폰 / 중개이용료 / ...)
  row 2~3: 세부 컬럼 헤더 (병합 셀)
  row 4~: 주문 단위 데이터 (1주문 = 1행)

전체 43컬럼 인덱스 매핑은 ParsedOrderFee 의 from_row() 참조.
주의: 컬럼 위치 기반 매핑 — 쿠팡이츠가 헤더 텍스트 표기를 바꿔도 인덱스가 같으면 동작.
"""
from __future__ import annotations

import datetime
import io
import logging
import re
from dataclasses import dataclass, field
from typing import Optional, Union

import openpyxl

log = logging.getLogger("coupang_eats.excel")


class ExcelParseError(Exception):
    """엑셀 파일 파싱 실패 (포맷 변경/손상/잘못된 파일)."""


# ──────────────────────────────────────────────────────────────────────────
# Data classes
# ──────────────────────────────────────────────────────────────────────────


@dataclass
class ParsedOrderFee:
    """주문 1건의 fee breakdown — 엑셀 43컬럼을 1:1 매핑."""

    # ── 주문정보 (0~8)
    order_date: datetime.date
    ordered_at: Optional[datetime.datetime]
    order_id: str
    order_type: Optional[str]                    # 배달 / 포장
    items_summary: Optional[str]                 # "모둠 김밥, 꼬마김밥"
    brand: Optional[str]
    shop_name: Optional[str]
    payment_method: Optional[str]                # 신용카드 / 현금 / ...
    transaction_type: Optional[str]              # 결제 / 취소

    # ── 매출액 (9~11)
    total_amount: int = 0          # 총금액 (쿠폰포함 정가)
    order_amount: int = 0          # 주문금액 (쿠폰차감 후)
    payment_amount: int = 0        # 결제금액 (고객 실결제)

    # ── 쿠폰 (12~13)
    coupon_coupang: int = 0        # 쿠팡부담 쿠폰
    coupon_store: int = 0          # 상점부담 쿠폰

    # ── 중개이용료 (14~16)
    brokerage_before_basic: int = 0    # 산정전 기본요금
    brokerage_before_promo: int = 0    # 산정전 프로모션
    brokerage_final: int = 0           # 산정후 (실차감) ★ P/L

    # ── 결제대행사 수수료 (17~18)
    payment_fee_basic: int = 0
    payment_fee_promo: int = 0

    # ── 배달비 (19~25)
    delivery_before_basic: int = 0
    delivery_before_promo: int = 0
    delivery_final: int = 0            # 산정후 ★ P/L
    delivery_only: int = 0             # 배달전용
    food_only: int = 0                 # 음식전용
    customer_delivery_fee: int = 0     # 고객부담배달비
    customer_delivery_fee_total: int = 0

    # ── 서비스이용료(멤버십) (26~33)
    service_before_disposable_cup: int = 0
    service_before_supply: int = 0
    service_before_vat: int = 0
    service_before_total: int = 0
    service_after_disposable_cup: int = 0
    service_after_supply: int = 0
    service_after_vat: int = 0
    service_after_total: int = 0       # 산정후 ★ P/L

    # ── 광고비 (34~36)
    ad_supply: int = 0
    ad_vat: int = 0
    ad_total: int = 0                  # ★ P/L

    # ── 정산금액 (37~39)
    settle_before_basic: int = 0
    settle_before_promo: int = 0
    settle_final: int = 0              # 산정후 (실수령) ★ 검증용

    # ── 프로모션 혜택 / 환급액 (40~42)
    extra_col_40: int = 0              # 미상 (헤더 비어있음 — 보존)
    promotion_benefit: int = 0
    refund_amount: int = 0

    # ── 메타
    raw_row: tuple = field(default_factory=tuple)   # 원본 43셀 tuple 보존

    # ── 도출 (취소 여부)
    @property
    def cancelled(self) -> bool:
        return (self.transaction_type or "").strip() == "취소"

    @classmethod
    def from_row(cls, row: tuple, idx_map: dict, *,
                 line_no: Optional[int] = None) -> "ParsedOrderFee":
        """헤더 매핑 dict 를 사용해 row 에서 ParsedOrderFee 생성.

        idx_map: build_column_index_map(ws) 의 결과 — 필드명 → 컬럼 인덱스.
        쿠팡이츠 엑셀 포맷이 시간 따라 변해도 (43→49컬럼) 헤더 기반이라 안정.
        """
        def _get_int(field: str) -> int:
            idx = idx_map.get(field)
            if idx is None or idx >= len(row):
                return 0
            return _to_int(row[idx])

        def _get_str(field: str) -> Optional[str]:
            idx = idx_map.get(field)
            if idx is None or idx >= len(row):
                return None
            return _to_str(row[idx])

        date_idx = idx_map.get('order_date')
        if date_idx is None or date_idx >= len(row):
            raise ExcelParseError(f"order_date 컬럼 매핑 실패 line={line_no}")
        order_date = _to_date(row[date_idx])
        if not order_date:
            raise ExcelParseError(f"order_date 파싱 실패: {row[date_idx]!r} line={line_no}")

        time_idx = idx_map.get('ordered_at')
        ordered_at = _to_datetime(row[time_idx]) if time_idx is not None and time_idx < len(row) else None

        order_id_idx = idx_map.get('order_id')
        order_id = str(row[order_id_idx] or "").strip() if order_id_idx is not None and order_id_idx < len(row) else ""
        if not order_id:
            raise ExcelParseError(f"order_id 비어있음 line={line_no}")

        return cls(
            order_date=order_date,
            ordered_at=ordered_at,
            order_id=order_id,
            order_type=_get_str('order_type'),
            items_summary=_get_str('items_summary'),
            brand=_get_str('brand'),
            shop_name=_get_str('shop_name'),
            payment_method=_get_str('payment_method'),
            transaction_type=_get_str('transaction_type'),
            total_amount=_get_int('total_amount'),
            order_amount=_get_int('order_amount'),
            payment_amount=_get_int('payment_amount'),
            coupon_coupang=_get_int('coupon_coupang'),
            coupon_store=_get_int('coupon_store'),
            brokerage_before_basic=_get_int('brokerage_before_basic'),
            brokerage_before_promo=_get_int('brokerage_before_promo'),
            brokerage_final=_get_int('brokerage_final'),
            payment_fee_basic=_get_int('payment_fee_basic'),
            payment_fee_promo=_get_int('payment_fee_promo'),
            delivery_before_basic=_get_int('delivery_before_basic'),
            delivery_before_promo=_get_int('delivery_before_promo'),
            delivery_final=_get_int('delivery_final'),
            delivery_only=_get_int('delivery_only'),
            food_only=_get_int('food_only'),
            customer_delivery_fee=_get_int('customer_delivery_fee'),
            customer_delivery_fee_total=_get_int('customer_delivery_fee_total'),
            service_before_disposable_cup=_get_int('service_before_disposable_cup'),
            service_before_supply=_get_int('service_before_supply'),
            service_before_vat=_get_int('service_before_vat'),
            service_before_total=_get_int('service_before_total'),
            service_after_disposable_cup=_get_int('service_after_disposable_cup'),
            service_after_supply=_get_int('service_after_supply'),
            service_after_vat=_get_int('service_after_vat'),
            service_after_total=_get_int('service_after_total'),
            ad_supply=_get_int('ad_supply'),
            ad_vat=_get_int('ad_vat'),
            ad_total=_get_int('ad_total'),
            settle_before_basic=_get_int('settle_before_basic'),
            settle_before_promo=_get_int('settle_before_promo'),
            settle_final=_get_int('settle_final'),
            extra_col_40=0,        # legacy — 더 이상 사용 안 함
            promotion_benefit=_get_int('promotion_benefit'),
            refund_amount=_get_int('refund_amount'),
            raw_row=tuple(row),
        )


@dataclass
class ParseReport:
    """파싱 결과 메타."""
    sheet_name: str
    total_rows: int = 0
    parsed_count: int = 0
    skipped_count: int = 0
    skip_reasons: list[str] = field(default_factory=list)
    period_start: Optional[datetime.date] = None
    period_end: Optional[datetime.date] = None
    orders: list[ParsedOrderFee] = field(default_factory=list)


# ──────────────────────────────────────────────────────────────────────────
# 변환 헬퍼
# ──────────────────────────────────────────────────────────────────────────


_DATE_FORMATS = ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d")


def _to_date(value) -> Optional[datetime.date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    s = str(value).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_datetime(value) -> Optional[datetime.datetime]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, datetime.date):
        return datetime.datetime.combine(value, datetime.time.min)
    s = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S",
                "%Y.%m.%d %H:%M:%S", "%Y.%m.%d %H:%M"):
        try:
            return datetime.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _to_str(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _to_int(value) -> int:
    """엑셀 셀 값 → int (음수/0/-0/None/문자열 다 안전).

    예: -0.0 → 0, 16600.0 → 16600, '1,234' → 1234, None → 0
    """
    if value is None or value == "":
        return 0
    if isinstance(value, bool):  # python에서 True == 1 이라 명시 차단
        return 0
    if isinstance(value, (int, float)):
        try:
            return int(round(float(value)))
        except (OverflowError, ValueError):
            return 0
    s = str(value).strip().replace(",", "").replace(" ", "")
    if not s:
        return 0
    try:
        return int(round(float(s)))
    except (ValueError, OverflowError):
        return 0


# ──────────────────────────────────────────────────────────────────────────
# 파서 본체
# ──────────────────────────────────────────────────────────────────────────

# 시트명 패턴: '2026-01_1', '2026-01' 등
_SHEET_NAME_RE = re.compile(r"^(\d{4})-(\d{2})(?:_\d+)?$")


# 필드명 → 헤더 path tuple(s). 여러 path 면 첫 매칭 우선.
# 쿠팡이츠가 새 컬럼 추가하더라도 헤더 텍스트 매핑이라 안정.
# 2026-04 부터 "광고 프로모션" + "최종 광고비" 컬럼 그룹이 추가됨 (43→49컬럼).
_FIELD_LABEL_PATHS: dict[str, tuple] = {
    'order_date':           (('주문정보', '일자'),),
    'ordered_at':           (('주문정보', '시간'),),
    'order_id':             (('주문정보', '주문번호'),),
    'order_type':           (('주문정보', '유형'),),
    'items_summary':        (('주문정보', '상세내역'),),
    'brand':                (('주문정보', '브랜드명'),),
    'shop_name':            (('주문정보', '상점명'),),
    'payment_method':       (('주문정보', '결제방식'),),
    'transaction_type':     (('주문정보', '거래유형'),),

    'total_amount':         (('매출액', '총금액'),),
    'order_amount':         (('매출액', '주문금액'),),
    'payment_amount':       (('매출액', '결제금액'),),

    'coupon_coupang':       (('쿠폰', '쿠팡부담'),),
    'coupon_store':         (('쿠폰', '상점부담'),),

    'brokerage_before_basic': (('중개이용료', '산정전', '기본요금'),),
    'brokerage_before_promo': (('중개이용료', '산정전', '프로모션'),),
    'brokerage_final':        (('중개이용료', '산정후'),),

    'payment_fee_basic':    (('결제대행사 수수료', '기본요금'),),
    'payment_fee_promo':    (('결제대행사 수수료', '프로모션'),),

    'delivery_before_basic': (('배달비', '산정전', '기본요금'),),
    'delivery_before_promo': (('배달비', '산정전', '프로모션'),),
    'delivery_final':        (('배달비', '산정후'),),

    'delivery_only':         (('즉시할인', '배달전용'),),
    'food_only':             (('즉시할인', '음식전용'),),

    'customer_delivery_fee':       (('기타', '고객부담배달비'),),
    'customer_delivery_fee_total': (('기타', '고객부담배달비(총액)'),),

    'service_before_disposable_cup': (('서비스이용료', '산정전', '일회용컵'),),
    'service_before_supply':         (('서비스이용료', '산정전', '공급가액'),),
    'service_before_vat':            (('서비스이용료', '산정전', '부가세액'),),
    'service_before_total':          (('서비스이용료', '산정전', '총액'),),
    'service_after_disposable_cup':  (('서비스이용료', '산정후', '일회용컵'),),
    'service_after_supply':          (('서비스이용료', '산정후', '공급가액'),),
    'service_after_vat':             (('서비스이용료', '산정후', '부가세액'),),
    'service_after_total':           (('서비스이용료', '산정후', '총액'),),

    # 광고비: 4월 이후 '최종 광고비' (광고비 + 광고 프로모션 합) 우선, 없으면 (1월) '광고비'
    'ad_supply': (('최종 광고비', '공급가액'), ('광고비', '공급가액')),
    'ad_vat':    (('최종 광고비', '부가세액'), ('광고비', '부가세액')),
    'ad_total':  (('최종 광고비', '총액'),     ('광고비', '총액')),

    'settle_before_basic':  (('정산금액', '산정전', '기본요금'),),
    'settle_before_promo':  (('정산금액', '산정전', '프로모션'),),
    'settle_final':         (('정산금액', '산정후'),),

    'promotion_benefit':    (('프로모션 혜택',),),
    'refund_amount':        (('환급액',),),
}


def _build_column_index_map(ws) -> tuple[dict[str, int], list[tuple]]:
    """row 1+2+3 헤더 → 필드명 → 컬럼 인덱스 매핑.

    헤더 셀의 None 은 직전 non-None 값으로 forward-fill (병합 셀 처리).
    단 row 3 의 None 은 그대로 유지 (산정후 같은 단일 컬럼).

    Returns:
        (idx_map, paths) — idx_map: 필드 → 컬럼 인덱스
                          paths:   디버그용 각 컬럼의 path tuple
    """
    def _cell_str(c):
        if c is None or c == "":
            return None
        s = str(c).strip()
        return s or None

    r1 = [_cell_str(c.value) for c in ws[1]]
    r2 = [_cell_str(c.value) for c in ws[2]]
    r3 = [_cell_str(c.value) for c in ws[3]]

    def _ffill(arr):
        out = []
        last = None
        for v in arr:
            if v is not None:
                last = v
            out.append(last)
        return out

    # r1, r2 는 그룹 헤더라 병합 → ffill
    # r3 는 detail 헤더 — None 이면 진짜 없음 (ffill 안 함)
    r1_ff = _ffill(r1)
    r2_ff = _ffill(r2)

    # 컬럼별 path tuple 생성
    paths: list[tuple] = []
    for i in range(ws.max_column):
        parts = []
        if r1_ff[i]: parts.append(r1_ff[i])
        if r2_ff[i]: parts.append(r2_ff[i])
        if r3[i]:    parts.append(r3[i])
        paths.append(tuple(parts))

    # 필드명 → 컬럼 인덱스 매핑
    idx_map: dict[str, int] = {}
    for field, candidate_paths in _FIELD_LABEL_PATHS.items():
        for target in candidate_paths:
            for i, p in enumerate(paths):
                if p == target:
                    idx_map[field] = i
                    break
            if field in idx_map:
                break

    return idx_map, paths


def parse_sales_order_excel(source: Union[str, bytes, io.BytesIO],
                            *,
                            expected_year_month: Optional[str] = None
                            ) -> ParseReport:
    """월별 매출내역서 엑셀 → ParseReport.

    Args:
        source: 파일 경로(str) 또는 바이트(bytes) 또는 BytesIO.
        expected_year_month: 'YYYY-MM' — 시트명/날짜와 검증. 불일치 시 경고 로그.

    Returns:
        ParseReport — 주문 list + 메타. 한 행 파싱 실패해도 다음 행 계속 (errors 수집).
    """
    if isinstance(source, bytes):
        source = io.BytesIO(source)

    try:
        # read_only=False 인 이유: 셀 병합·이름 조회 등을 위해서.
        # 87KB 정도 작은 파일이라 메모리 OK.
        wb = openpyxl.load_workbook(source, data_only=True)
    except Exception as e:
        raise ExcelParseError(f"엑셀 로드 실패: {e}") from e

    if not wb.sheetnames:
        raise ExcelParseError("시트가 없습니다.")

    # 첫 번째 시트 사용 (쿠팡이츠는 보통 'YYYY-MM_1' 단일 시트)
    ws = wb[wb.sheetnames[0]]
    sheet_name = ws.title

    # 컬럼 수 sanity (43=1월 포맷 / 49=2026-04 이후 포맷 / 추후 더 늘어날 수 있음)
    if ws.max_column < 30:
        raise ExcelParseError(
            f"컬럼 수 {ws.max_column} 너무 적음 — 포맷 깨짐 의심. sheet={sheet_name}"
        )

    # 헤더 → 필드 인덱스 매핑 (포맷 변경에도 안정)
    idx_map, paths = _build_column_index_map(ws)

    # 필수 필드 확인
    required = ('order_date', 'order_id', 'total_amount', 'brokerage_final',
                'delivery_final', 'ad_total', 'settle_final')
    missing = [f for f in required if f not in idx_map]
    if missing:
        # 디버그: 처음 60개 path 출력
        path_dump = "\n".join(f"  col[{i:2d}]: {' / '.join(p) if p else '(empty)'}"
                              for i, p in enumerate(paths[:60]))
        raise ExcelParseError(
            f"필수 컬럼 매핑 실패 {missing} — sheet={sheet_name} max_col={ws.max_column}\n"
            f"감지된 헤더:\n{path_dump}"
        )

    log.info(
        "excel column map sheet=%s max_col=%d mapped=%d/%d",
        sheet_name, ws.max_column, len(idx_map), len(_FIELD_LABEL_PATHS),
    )

    # 시트명에서 연-월 추출 (정합성 검증용)
    m = _SHEET_NAME_RE.match(sheet_name)
    parsed_ym = f"{m.group(1)}-{m.group(2)}" if m else None
    if expected_year_month and parsed_ym and parsed_ym != expected_year_month:
        log.warning(
            "expected_year_month=%s 이지만 sheet=%s (parsed=%s) — 진행하지만 의심됨",
            expected_year_month, sheet_name, parsed_ym,
        )

    report = ParseReport(sheet_name=sheet_name, total_rows=max(0, ws.max_row - 3))

    # row 4 부터 데이터. 빈 행이 끝까지 갈 수 있어 break 처리.
    for idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        # 완전 빈 행 (모든 셀 None) → 데이터 끝
        if all(c is None or c == "" for c in row[:9]):
            continue
        try:
            order = ParsedOrderFee.from_row(row, idx_map, line_no=idx)
        except ExcelParseError as e:
            report.skipped_count += 1
            reason = f"line={idx} {e}"
            report.skip_reasons.append(reason)
            log.warning("excel row skipped: %s", reason)
            continue
        report.orders.append(order)
        report.parsed_count += 1

        # 기간 추적
        if report.period_start is None or order.order_date < report.period_start:
            report.period_start = order.order_date
        if report.period_end is None or order.order_date > report.period_end:
            report.period_end = order.order_date

    log.info(
        "excel parsed sheet=%s rows=%d parsed=%d skipped=%d period=%s~%s",
        sheet_name, report.total_rows, report.parsed_count, report.skipped_count,
        report.period_start, report.period_end,
    )
    return report


# ──────────────────────────────────────────────────────────────────────────
# 일자별 집계 — settlement.fee_* 채움용
# ──────────────────────────────────────────────────────────────────────────


@dataclass
class DailyFeeAggregate:
    """일자별 fee 집계 — CoupangEatsSettlement.fee_* 컬럼에 적재.

    취소(transaction_type='취소') 주문 제외.
    """
    order_date: datetime.date
    order_count: int = 0
    cancelled_count: int = 0

    total_amount: int = 0          # 총금액 합
    order_amount: int = 0          # 주문금액 합
    payment_amount: int = 0        # 결제금액 합

    fee_brokerage: int = 0         # 중개수수료 산정후 합 ★
    fee_payment: int = 0           # 결제수수료 기본+프로모션 합 ★
    fee_delivery: int = 0          # 배달비 산정후 합 ★
    fee_advertising: int = 0       # 광고비 총액 합 ★
    fee_membership: int = 0        # 서비스이용료 산정후 총액 합 ★

    coupon_store: int = 0          # 상점부담 쿠폰 합 (P/L에 비용으로 잡힐 수 있음)
    settle_final: int = 0          # 정산금액 산정후 합 (검증용 = 정산 amount 와 일치)
    refund_amount: int = 0
    promotion_benefit: int = 0


def aggregate_by_date(orders: list[ParsedOrderFee]
                      ) -> dict[datetime.date, DailyFeeAggregate]:
    """주문 list → 일자별 집계 dict.

    취소 건은 fee 합산에서 제외. 단 cancelled_count 카운트만 함.
    """
    out: dict[datetime.date, DailyFeeAggregate] = {}
    for o in orders:
        agg = out.setdefault(o.order_date,
                             DailyFeeAggregate(order_date=o.order_date))
        if o.cancelled:
            agg.cancelled_count += 1
            continue
        agg.order_count += 1
        agg.total_amount += o.total_amount
        agg.order_amount += o.order_amount
        agg.payment_amount += o.payment_amount
        agg.fee_brokerage += o.brokerage_final
        agg.fee_payment += o.payment_fee_basic + o.payment_fee_promo
        agg.fee_delivery += o.delivery_final
        agg.fee_advertising += o.ad_total
        agg.fee_membership += o.service_after_total
        agg.coupon_store += o.coupon_store
        agg.settle_final += o.settle_final
        agg.refund_amount += o.refund_amount
        agg.promotion_benefit += o.promotion_benefit
    return out
