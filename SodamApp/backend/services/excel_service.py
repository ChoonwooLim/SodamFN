import pandas as pd
import os
import datetime

# Hardcoded path for now, can be moved to config
EXCEL_PATH = r"C:\WORK\SodamFN\2025실소득분석\소담김밥손익계산서(9~12).xlsx"

class ExcelService:
    def __init__(self, file_path=None):
        self.file_path = file_path
        # Only check existence if a real path is provided (not None and not "dummy_path")
        if self.file_path and self.file_path != "dummy_path" and not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Excel file not found at {self.file_path}")

    def get_monthly_summary(self):
        """
        Reads the '종합' sheet and extracts monthly Revenue, Cost, and Profit.
        Returns a list of dictionaries for the frontend graph.
        """
        try:
            # Read '종합' sheet, no header to rely on index
            df = pd.read_excel(self.file_path, sheet_name='종합', header=None)
            
            # Map columns to months based on our analysis
            # Col 4: Jul, 5: Aug, 6: Sep, 7: Oct, 8: Nov, 9: Dec
            # (0-indexed in pandas)
            month_map = {
                4: "7월", 5: "8월", 6: "9월", 7: "10월", 8: "11월", 9: "12월"
            }
            
            # Row indices (observed from analysis)
            # Revenue: Row 2 (Index 2) -> "수입 매장 매출" + others. 
            # Actually Row 7 (Index 7) is "합 계" of Revenue?
            # Let's look at preview again.
            # Row 7: 합 계 (Revenue Total)
            # Row 17: 합 계 (Expense Total)
            # Row 18: 영업이익 (Profit)
            
            revenue_row_idx = 7
            expense_row_idx = 17
            profit_row_idx = 18
            
            summary_data = []
            
            for col_idx, month_name in month_map.items():
                revenue = df.iloc[revenue_row_idx, col_idx]
                expense = df.iloc[expense_row_idx, col_idx]
                profit = df.iloc[profit_row_idx, col_idx]
                
                # Handle NaNs
                revenue = float(revenue) if pd.notna(revenue) else 0
                expense = float(expense) if pd.notna(expense) else 0
                profit = float(profit) if pd.notna(profit) else 0
                
                summary_data.append({
                    "month": month_name,
                    "revenue": int(revenue),
                    "expense": int(expense),
                    "profit": int(profit),
                    "margin": round((profit / revenue * 100), 1) if revenue > 0 else 0
                })
                
            return summary_data

        except Exception as e:
            print(f"Error reading Excel: {e}")
            return []

    def get_revenue_breakdown(self):
        """
        Extracts revenue sources (Store, Coupang, Baemin, etc.) for the latest month (Dec).
        """
        try:
            df = pd.read_excel(self.file_path, sheet_name='종합', header=None)
            
            # Row indices for channels (0-indexed from inspection)
            # Row 2: 매장 매출
            # Row 3: 쿠팡 정산금
            # Row 4: 배민 정산금
            # Row 5: 요기요 정산금
            # Row 6: 땡겨요 정산금
            channels = {
                "매장": 2,
                "쿠팡": 3,
                "배민": 4,
                "요기요": 5,
                "땡겨요": 6
            }
            
            # Target Month Column: 9 (December)
            target_col = 9 
            
            result = []
            for name, row_idx in channels.items():
                val = df.iloc[row_idx, target_col]
                val = int(val) if pd.notna(val) else 0
                result.append({"name": name, "value": val})
                
            return result
        except Exception as e:
            print(f"Error reading Revenue Breakdown: {e}")
            return []

    def get_top_expenses(self, month: int = 12):
        """
        Extracts top vendors by cost for a specific month.
        Default to December (12).
        
        Also joins with 'Vendor Info' sheet to get Item details.
        """
        try:
            # 1. Get Top Expenses from Cost Sheet
            sheet_name = f"{month}월비용"
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=None)
            
            # Vendor Name: Col 0 (A)
            # Total Amount: Col 31 (AF) - Last column based on inspection
            
            # Slice from row 2 (skipping headers)
            vendors = df.iloc[2:, 0] 
            amounts = df.iloc[2:, 31] 
            
            data = pd.DataFrame({"vendor": vendors, "amount": amounts})
            
            # Clean data
            data = data.dropna(subset=['amount'])
            # Ensure amount is numeric
            data['amount'] = pd.to_numeric(data['amount'], errors='coerce').fillna(0)
            data = data.dropna(subset=['vendor']) # Drop if vendor is NaN
            data = data[data['amount'] > 0]
            
            # Group by Vendor
            grouped = data.groupby('vendor')['amount'].sum().reset_index()
            grouped = grouped.sort_values('amount', ascending=False)
            top_list = grouped.head(5).to_dict('records')

            # 2. Get Item Info from 'Vendor Info' Sheet
            vendor_info = self.get_vendors() # Returns dict {name: item}
            
            # 3. Merge Info
            result = []
            for item in top_list:
                v_name = item['vendor']
                v_item = vendor_info.get(v_name, "")
                result.append({
                    "vendor": v_name,
                    "amount": int(item['amount']),
                    "item": v_item
                })
                
            return result
            
        except Exception as e:
            print(f"Error reading Top Expenses: {e}")
            return []

    def get_vendors(self):
        """
        Reads '거래처정보' sheet and returns a dict mapping Vendor -> Item.
        If sheet doesn't exist, returns empty dict.
        """
        try:
            wb = load_workbook(self.file_path)
            if '거래처정보' not in wb.sheetnames:
                return {}
            
            df = pd.read_excel(self.file_path, sheet_name='거래처정보')
            # Expecting columns: ['Vendor', 'Item']
            if df.empty:
                return {}
            
            return pd.Series(df.Item.values, index=df.Vendor).to_dict()
        except Exception as e:
            # print(f"Error reading Vendor Info: {e}")
            return {}

    def sync_vendors(self):
        """
        Scans all monthly sheets, finds unique vendors, and adds missing ones to '거래처정보' sheet.
        """
        try:
            wb = load_workbook(self.file_path)
            
            # 1. Collect all vendors from monthly sheets
            all_vendors = set()
            months = [7, 8, 9, 10, 11, 12]
            
            for m in months:
                sheet_name = f"{m}월비용"
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # Scan Column A (idx 1), skipping first 2 rows
                    for row in ws.iter_rows(min_row=3, min_col=1, max_col=1, values_only=True):
                        val = row[0]
                        if val and isinstance(val, str):
                            all_vendors.add(val)
            
            # 2. Initialize or Load '거래처정보' sheet
            if '거래처정보' not in wb.sheetnames:
                ws_info = wb.create_sheet('거래처정보')
                ws_info.append(['Vendor', 'Item']) # Header
                existing_vendors = set()
            else:
                ws_info = wb['거래처정보']
                # Read existing
                existing_vendors = set()
                for row in ws_info.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
                    if row[0]: existing_vendors.add(row[0])
            
            # 3. Add new vendors
            new_vendors = all_vendors - existing_vendors
            for v in new_vendors:
                ws_info.append([v, '']) # Default empty item
                
            if new_vendors:
                wb.save(self.file_path)
                return len(new_vendors)
            return 0
            
        except Exception as e:
            print(f"Error syncing vendors: {e}")
            return 0

    def update_vendor_item(self, vendor_name, new_item):
        """
        Updates the 'Item' column for a specific vendor in '거래처정보' sheet.
        """
        try:
            wb = load_workbook(self.file_path)
            if '거래처정보' not in wb.sheetnames:
                return False
            
            ws = wb['거래처정보']
            found = False
            
            # Search for vendor
            for row in ws.iter_rows(min_row=2, max_col=2):
                cell_vendor = row[0]
                cell_item = row[1]
                
                if cell_vendor.value == vendor_name:
                    cell_item.value = new_item
                    found = True
                    break
            
            if found:
                wb.save(self.file_path)
                return True
            return False
            
        except Exception as e:
            print(f"Error updating vendor item: {e}")
            return False

    def add_expense(self, date_str: str, item: str, amount: int, category: str = "기타"):
        """
        Adds an expense record to the appropriate monthly sheet.
        date_str: "YYYY-MM-DD"
        """
        try:
            # Parse date to find month
            dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")
            month = dt.month
            
            # Map month to sheet name (Need to match Excel sheet names like '9월비용')
            # The file has '7월비용' ... '12월비용'
            sheet_name = f"{month}월비용"
            
            # Helper to check if sheet exists
            xl = pd.ExcelFile(self.file_path)
            if sheet_name not in xl.sheet_names:
                # Fallback or error?
                # Maybe the file only has 9-12?
                # Let's check available sheets
                if sheet_name not in xl.sheet_names:
                    return {"status": "error", "message": f"Sheet {sheet_name} not found"}
            
            # Load the workbook using openpyxl for editing
            from openpyxl import load_workbook
            wb = load_workbook(self.file_path)
            ws = wb[sheet_name]
            
            # Find first empty row (assuming standard list format)
            # We should append to the bottom.
            # Assuming columns: Date | Item | Amount | Category ...
            # Need to know the column structure of cost sheets.
            # For now, let's append to the end.
            ws.append([date_str, item, amount, category])
            
            wb.save(self.file_path)
            return {"status": "success", "message": f"Added to {sheet_name}"}
            
        except Exception as e:
            print(f"Error adding expense: {e}")
            return {"status": "error", "message": str(e)}
    def parse_upload(self, file_contents: bytes):
        """
        Smart parser that auto-detects card company format and parses accordingly.
        Supports: 롯데카드, 신한카드, 삼성카드, KB국민카드, 현대카드, 은행 송금내역, 일반 형식
        Supports both .xls and .xlsx formats
        """
        import io
        try:
            # Helper function to read excel with fallback engines
            def read_excel_safe(content, **kwargs):
                """Try different engines for .xls/.xlsx compatibility and handle HTML masquerading as Excel"""
                error_list = []
                
                # 1. Try standard Excel engines
                engines = ['openpyxl', 'xlrd', None]
                for engine in engines:
                    try:
                        if engine:
                            return pd.read_excel(io.BytesIO(content), engine=engine, **kwargs)
                        else:
                            return pd.read_excel(io.BytesIO(content), **kwargs)
                    except Exception as e:
                        error_list.append(f"Engine {engine}: {str(e)}")
                        continue
                
                # 2. Try parsing as HTML (common in Korean finance for .xls files)
                try:
                    # read_html does not support 'nrows', remove it if present
                    html_kwargs = kwargs.copy()
                    if 'nrows' in html_kwargs:
                        del html_kwargs['nrows']
                        
                    dfs = pd.read_html(io.BytesIO(content), **html_kwargs)
                    if dfs:
                        # Return the dataframe with the most rows/columns roughly
                        best_df = max(dfs, key=lambda x: x.size)
                        return best_df
                except Exception as e:
                    error_list.append(f"HTML Parser: {str(e)}")
                
                # If all failed
                raise ValueError(f"Failed to read file. Errors: {'; '.join(error_list)}")
            
            # Read the file
            # Note: We skip the 'preview' step with nrows because read_html doesn't support it 
            # and it causes issues. We'll read the whole file and find headers in-memory.
            try:
                df = read_excel_safe(file_contents)
            except Exception as e:
                 return {"status": "error", "message": f"파일 읽기 실패: {str(e)}"}

            # Post-load Header Detection
            # If the first row doesn't look like a header, scan specifically for headers
            cols = [str(c).strip() for c in df.columns.tolist()]
            
            # Keywords to identify header row
            header_keywords = ['이용일자', '승인일자', '거래일자', '날짜', 'Date', '일자']
            
            found_header_idx = -1
            
            # Check if current columns are already headers
            if any(k in str(c) for c in cols for k in header_keywords):
                found_header_idx = -2 # Already good
            else:
                # Scan first 20 rows for header
                for i, row in df.head(20).iterrows():
                    row_text = ' '.join([str(v) for v in row.values if pd.notna(v)])
                    if any(k in row_text for k in header_keywords):
                        found_header_idx = i
                        break
            
            # Apply new header if found
            if found_header_idx >= 0:
                # Set row i as header
                new_header = df.iloc[found_header_idx]
                df = df.iloc[found_header_idx + 1:]
                df.columns = new_header
                cols = [str(c).strip() for c in df.columns.tolist()]
                
            # Now proceed with detection
            cols_lower = [c.lower() for c in cols]
            
            # Detect format and map columns
            format_detected = "unknown"
            date_col = None
            item_col = None
            amount_col = None
            category_col = None
            
            # Helper to find column by keywords
            def find_col(keywords):
                for i, c in enumerate(cols):
                    if any(k in c for k in keywords):
                        return cols[i]
                return None
            
            # Pattern 1: 롯데카드/신한카드 형식 (이용일자, 이용가맹점, 이용금액)
            if find_col(['이용일자', '승인일자', '결제일자']):
                format_detected = "card_statement"
                date_col = find_col(['이용일자', '승인일자', '결제일자', '거래일자'])
                item_col = find_col(['이용가맹점', '가맹점명', '가맹점', '사용처', '이용처'])
                amount_col = find_col(['이용금액', '결제금액', '승인금액', '사용금액', '금액'])
                category_col = find_col(['업종', '분류', '카테고리'])
            
            # Pattern 1.5: 신한은행/국민은행 송금내역 형식 (거래일자, 출금(원), 내용)
            elif find_col(['거래일자']) and find_col(['출금']):
                format_detected = "bank_transfer"
                date_col = find_col(['거래일자'])
                # Prioritize '내용' (counterparty name) over '적요' (transfer method like '모바일')
                item_col = find_col(['내용', '거래내용', '메모']) or find_col(['적요'])
                amount_col = find_col(['출금', '출금(원)', '출금액', '이체금액'])
                category_col = None  # Bank statements usually don't have category
                
            # Pattern 2: 일반 지출 형식 (날짜, 항목, 금액)
            elif find_col(['날짜', 'Date', '일자']):
                format_detected = "standard"
                date_col = find_col(['날짜', 'Date', '일자'])
                item_col = find_col(['항목', 'Item', '내역', '사용처', '적요'])
                amount_col = find_col(['금액', 'Amount', '비용', '지출'])
                category_col = find_col(['분류', 'Category', '카테고리'])
            
            # Pattern 3: Fallback - use first few columns
            else:
                format_detected = "auto"
                # Try to find date-like column
                for c in cols:
                    if any(x in c for x in ['일', 'date', 'Date', '날']):
                        date_col = c
                        break
                if not date_col and len(cols) > 0:
                    date_col = cols[0]
                    
                # Amount column - look for numbers in column name or data
                for c in cols:
                    if any(x in c for x in ['금액', '원', 'amount', 'Amount', '비용']):
                        amount_col = c
                        break
                if not amount_col and len(cols) > 1:
                    # Try to find a numeric column
                    for c in cols[1:]:
                        try:
                            if df[c].dtype in ['int64', 'float64']:
                                amount_col = c
                                break
                        except:
                            pass
                            
                item_col = find_col(['항목', '가맹점', '내역', '사용처', '적요', '상호'])
            
            if not date_col:
                return {"status": "error", "message": "날짜 컬럼을 찾을 수 없습니다. (이용일자/날짜/Date)"}
            if not amount_col:
                return {"status": "error", "message": "금액 컬럼을 찾을 수 없습니다. (이용금액/금액/Amount)"}
            
            results = []
            skipped = 0
            
            for idx, row in df.iterrows():
                try:
                    # Parse date
                    d_val = row[date_col]
                    if pd.isna(d_val):
                        skipped += 1
                        continue
                        
                    if isinstance(d_val, (datetime.datetime, datetime.date)):
                        d_str = d_val.strftime("%Y-%m-%d")
                    elif isinstance(d_val, str):
                        # Handle various date formats
                        d_clean = d_val.strip()
                        
                        # Korean format: "2026년 01월 31일" or "2026년01월31일"
                        if '년' in d_clean:
                            import re
                            match = re.search(r'(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일?', d_clean)
                            if match:
                                year, month, day = match.groups()
                                d_str = f"{year}-{int(month):02d}-{int(day):02d}"
                            else:
                                # Only year and month: "2026년 01월" - skip this row
                                skipped += 1
                                continue
                        else:
                            # Standard formats: 2026.01.31, 2026-01-31, 2026/01/31
                            d_clean = d_clean.replace('.', '-').replace('/', '-')
                            d_str = d_clean[:10]
                    else:
                        d_str = str(d_val)[:10]
                    
                    # Validate date format (must be YYYY-MM-DD with at least 10 chars)
                    if not d_str or len(d_str) < 8 or d_str in ['-', '--', '---']:
                        skipped += 1
                        continue
                    # Additional check: must contain digits
                    if not any(c.isdigit() for c in d_str):
                        skipped += 1
                        continue
                    
                    # Parse amount
                    amt = row[amount_col]
                    if pd.isna(amt):
                        skipped += 1
                        continue
                    
                    # Handle various number formats: 7,900 / 7900 / 7900.0
                    if isinstance(amt, (int, float)):
                        amt = int(abs(amt))
                    else:
                        amt_str = str(amt).replace(',', '').replace('원', '').strip()
                        try:
                            amt = int(abs(float(amt_str)))
                        except:
                            skipped += 1
                            continue
                    
                    if amt == 0:
                        skipped += 1
                        continue
                    
                    # Parse item/vendor
                    item = "미지정"
                    if item_col and pd.notna(row.get(item_col)):
                        item = str(row[item_col]).strip()
                    
                    # Parse category
                    cat = "기타"
                    if category_col and pd.notna(row.get(category_col)):
                        cat = str(row[category_col]).strip()
                    
                    results.append({
                        "date": d_str,
                        "item": item,
                        "amount": amt,
                        "category": cat
                    })
                    
                except Exception as row_e:
                    skipped += 1
                    continue
            
            if not results:
                return {"status": "error", "message": f"파싱된 데이터가 없습니다. (형식: {format_detected}, 스킵: {skipped}행)"}
                    
            return {
                "status": "success", 
                "data": results,
                "format": format_detected,
                "parsed": len(results),
                "skipped": skipped
            }
            
        except Exception as e:
            import traceback
            return {"status": "error", "message": f"Excel 파싱 오류: {str(e)}"}

    # --- Card company → vendor name mapping for revenue upload ---
    CARD_VENDOR_MAP = {
        '신한카드': '소담김밥 건대본점 신한카드',
        'KB카드': '소담김밥 건대본점 국민카드',
        'KB국민카드': '소담김밥 건대본점 국민카드',
        '비씨카드': '소담김밥 건대본점 BC카드',
        '현대카드': '소담김밥 건대본점 현대카드',
        '하나카드': '소담김밥 건대본점 하나카드',
        '하나구외환': '소담김밥 건대본점 하나카드',
        '삼성카드': '소담김밥 건대본점 삼성카드',
        '롯데카드': '소담김밥 건대본점 롯데카드',
        '신롯데카드': '소담김밥 건대본점 롯데카드',
        '우리카드': '소담김밥 건대본점 우리카드',
        '농협카드': '소담김밥 건대본점 농협카드',
        'NH카드': '소담김밥 건대본점 농협카드',
        'NH농협카드': '소담김밥 건대본점 농협카드',
        '카카오페이': '소담김밥 건대본점 카카오페이',
        # Pay services
        '제로페이': '소담김밥 건대본점 제로페이',
        '네이버페이': '소담김밥 건대본점 네이버페이',
        '애플페이': '소담김밥 건대본점 애플페이',
        '삼성페이': '소담김밥 건대본점 삼성페이',
        '페이코': '소담김밥 건대본점 페이코',
        'PAYCO': '소담김밥 건대본점 페이코',
    }
    CASH_VENDOR_NAME = '소담김밥 건대본점 현금매출'

    def parse_revenue_upload(self, file_contents: bytes, password: str = None):
        """
        Smart revenue upload parser.
        Detects file type and delegates to appropriate parser.
        Supports password-protected Excel files.
        
        Returns:
            dict with keys: status, file_type, data[], summary{}
            data items: { date, amount, vendor_name, note }
        """
        import io
        
        try:
            # ── Try to decrypt password-protected files ──
            COMMON_PASSWORDS = ['630730', '1234', '0000', '1111']
            passwords_to_try = []
            if password:
                passwords_to_try.append(password)
            passwords_to_try.extend(COMMON_PASSWORDS)
            
            decrypted = False
            is_encrypted = False
            try:
                import msoffcrypto
                office_file = msoffcrypto.OfficeFile(io.BytesIO(file_contents))
                if office_file.is_encrypted():
                    is_encrypted = True
                    for pwd in passwords_to_try:
                        try:
                            buf = io.BytesIO()
                            office_file = msoffcrypto.OfficeFile(io.BytesIO(file_contents))
                            office_file.load_key(password=pwd)
                            office_file.decrypt(buf)
                            file_contents = buf.getvalue()
                            decrypted = True
                            break
                        except Exception:
                            continue
                    if not decrypted:
                        return {"status": "password_required", "message": "비밀번호가 설정된 파일입니다. 비밀번호를 입력해주세요."}
            except ImportError:
                pass
            except Exception:
                pass

            # Try to read the file with different engines
            df = None
            for engine_name in ['openpyxl', 'xlrd', None]:
                try:
                    if engine_name:
                        df = pd.read_excel(io.BytesIO(file_contents), header=None, engine=engine_name)
                    else:
                        df = pd.read_excel(io.BytesIO(file_contents), header=None)
                    break
                except Exception:
                    continue
            
            if df is None:
                # Try HTML (Korean .xls files are often HTML)
                try:
                    dfs = pd.read_html(io.BytesIO(file_contents))
                    if dfs:
                        df = max(dfs, key=lambda x: x.size)
                except Exception:
                    pass
            
            if df is None:
                return {"status": "error", "message": "파일을 읽을 수 없습니다."}
            
            # --- Detect file type by scanning first few rows ---
            # Priority 1: Known specific formats (exact keyword match)
            first_rows_text = ''
            for i in range(min(15, len(df))):
                row_vals = [str(v) for v in df.iloc[i].tolist() if pd.notna(v)]
                first_rows_text += ' '.join(row_vals) + ' '
            
            if '매출일자' in first_rows_text and ('총매출' in first_rows_text or '순매출' in first_rows_text):
                return self._parse_pos_daily_revenue(df)
            elif '기간별 승인내역' in first_rows_text or ('No.' in first_rows_text and '카드사' in first_rows_text and '승인금액' in first_rows_text):
                return self._parse_card_detail_revenue(df)
            elif '월별 승인내역' in first_rows_text:
                return self._parse_card_summary_revenue(df)
            
            # Also try reading all sheets for Baemin detection
            # (Baemin files have 요약 + 상세 sheets, pandas only reads active/first sheet)
            all_sheets_data = None
            if is_encrypted or '정산명세서' in first_rows_text or ('주문중개' in first_rows_text and '입금금액' in first_rows_text):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(file_contents))
                    if '상세' in wb.sheetnames:
                        all_sheets_data = wb
                except Exception:
                    pass

            if '정산명세서' in first_rows_text or ('주문중개' in first_rows_text and '입금금액' in first_rows_text):
                return self._parse_baemin_settlement(df, workbook=all_sheets_data)

            # Priority 1.6: Yogiyo settlement (요기요 정산내역)
            if '상호명' in first_rows_text and '주문번호' in first_rows_text and '주문일시' in first_rows_text:
                return self._parse_yogiyo_settlement(df)
            elif '정산내역' in first_rows_text and ('주문금액' in first_rows_text or '요타임딜할인' in first_rows_text):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(file_contents))
                    if '상세 거래내역' in wb.sheetnames:
                        detail_df = pd.read_excel(io.BytesIO(file_contents), sheet_name='상세 거래내역', header=None)
                        return self._parse_yogiyo_settlement(detail_df)
                except Exception:
                    pass

            # Priority 1.6.5: 땡겨요 settlement (땡겨요 정산내역) — MUST be checked before Coupang Eats
            # Variants: 2월 format uses (D)차감금액/(E)정산금액, 1월 format uses (B)차감금액/(C)정산금액
            if '(A)주문결제' in first_rows_text and ('(D)차감금액' in first_rows_text or '(E)정산금액' in first_rows_text or ('(B)차감금액' in first_rows_text and '정산 내역' in first_rows_text)):
                return self._parse_ddangyo_settlement(df)

            # Priority 1.6.6: Coupang Eats settlement (쿠팡이츠 정산내역)
            if ('쿠팡' in first_rows_text and '정산금액' in first_rows_text) or ('주문정보' in first_rows_text and '정산금액' in first_rows_text) or ('브랜드명' in first_rows_text and '상점명' in first_rows_text and '주문금액' in first_rows_text and '정산금액' in first_rows_text):
                return self._parse_coupang_eats_settlement(df)

            # Priority 1.8: Delivery app packed format (쿠팡이츠 등)
            # Single-column format like: "1. 2026.02.27기본정산286,792원724,366원"
            if len(df.columns) <= 2 and ('기본정산' in first_rows_text or '인출' in first_rows_text):
                return self._parse_delivery_settlement(df)

            # Priority 1.9: Bank Deposit Statement (은행 입금내역) - Used to compute Card Fees
            if ('거래일시' in first_rows_text or '거래일자' in first_rows_text) and ('찾으신금액' in first_rows_text or '맡기신금액' in first_rows_text or '입금' in first_rows_text):
                res = self._parse_bank_deposit_for_card_fee(df)
                if res.get('status') == 'success':
                    return res
            
            # Priority 2: Universal pattern-based detection (any POS vendor)
            return self._parse_universal_revenue(df, file_contents)
        
        except Exception as e:
            import traceback
            return {"status": "error", "message": f"매출 파일 파싱 오류: {str(e)}"}

    def _parse_baemin_settlement(self, df, workbook=None):
        """
        Parse 배달의민족 정산명세서.
        Uses '상세' (detail) sheet for daily entries when available.
        Falls back to '요약' (summary) sheet if detail not available.
        """
        import re
        
        # Find year and month from title (요약 sheet)
        year, month = None, None
        for i in range(min(3, len(df))):
            row_text = ' '.join([str(v) for v in df.iloc[i].tolist() if pd.notna(v)])
            m = re.search(r'(\d{4})\s*년\s*(\d{1,2})\s*월', row_text)
            if m:
                year = int(m.group(1))
                month = int(m.group(2))
                break
        
        if not year or not month:
            return {"status": "error", "message": "정산명세서에서 년/월 정보를 찾을 수 없습니다."}
        
        import calendar
        import datetime
        
        data = []
        total_settlement = 0
        total_sales = 0
        total_fees = 0
        
        # ── Try detail sheet first ──
        if workbook and '상세' in workbook.sheetnames:
            ws = workbook['상세']
            
            # Find header row with 입금일, 입금 금액
            header_row = None
            col_map = {}  # column name → index
            
            for i in range(1, min(ws.max_row + 1, 10)):
                row_vals = [c.value for c in ws[i][:26]]
                for j, val in enumerate(row_vals):
                    cell_text = str(val).strip() if val else ''
                    if '입금일' in cell_text:
                        header_row = i
                        col_map['입금일'] = j
                    if '입금 금액' in cell_text or '입금금액' in cell_text:
                        col_map['입금금액'] = j
                    if '주문중개' in cell_text:
                        col_map['주문중개'] = j
                if header_row is not None:
                    break
            
            if header_row and '입금일' in col_map and '입금금액' in col_map:
                # Parse data rows — aggregate by date
                date_totals = {}  # date_str → {settlement, sales}
                start_date = datetime.date(year, month, 1)
                if month == 12:
                    end_date = datetime.date(year + 1, 1, 1)
                else:
                    end_date = datetime.date(year, month + 1, 1)
                
                for i in range(header_row + 1, ws.max_row + 1):
                    row_vals = [c.value for c in ws[i][:26]]
                    
                    # Get 입금일
                    raw_date = row_vals[col_map['입금일']]
                    if raw_date is None:
                        continue
                    
                    # Parse date
                    try:
                        if isinstance(raw_date, datetime.datetime):
                            d = raw_date.date()
                        elif isinstance(raw_date, datetime.date):
                            d = raw_date
                        else:
                            d = datetime.datetime.strptime(str(raw_date).strip(), "%Y-%m-%d").date()
                    except Exception:
                        continue
                    
                    # Only include dates within the target month
                    if d < start_date or d >= end_date:
                        continue
                    
                    # Get 입금금액
                    settlement = row_vals[col_map['입금금액']]
                    if settlement is None:
                        continue
                    try:
                        settlement = int(float(settlement))
                    except:
                        continue
                    
                    # Get 주문중개 (sales) if available
                    sales = 0
                    if '주문중개' in col_map:
                        sv = row_vals[col_map['주문중개']]
                        if sv is not None:
                            try:
                                sales = int(float(sv))
                            except:
                                pass
                    
                    date_str = d.strftime("%Y-%m-%d")
                    if date_str not in date_totals:
                        date_totals[date_str] = {"settlement": 0, "sales": 0}
                    date_totals[date_str]["settlement"] += settlement
                    date_totals[date_str]["sales"] += sales
                
                # Build data entries
                for date_str in sorted(date_totals.keys()):
                    dt = date_totals[date_str]
                    # Use sales (gross) as revenue amount; fallback to settlement if no sales data
                    revenue_amount = dt["sales"] if dt["sales"] > 0 else dt["settlement"]
                    data.append({
                        "date": date_str,
                        "amount": revenue_amount,
                        "vendor_name": "배달의민족",
                        "note": f"배민 정산",
                        "payment_type": "delivery",
                    })
                    total_settlement += dt["settlement"]
                    total_sales += dt["sales"]
        
        # ── Fallback to summary sheet ──
        if not data:
            header_row = None
            deposit_col = None
            order_col = None
            
            for i in range(len(df)):
                for j, val in enumerate(df.iloc[i].tolist()):
                    cell_text = str(val).strip() if pd.notna(val) else ''
                    if '입금금액' in cell_text:
                        header_row = i
                        deposit_col = j
                    if '주문중개' in cell_text:
                        order_col = j
                if header_row is not None:
                    break
            
            if header_row is not None and deposit_col is not None:
                for i in range(header_row + 1, len(df)):
                    val = df.iloc[i, deposit_col]
                    if pd.notna(val):
                        try:
                            total_settlement = int(float(val))
                        except:
                            continue
                        if order_col is not None and pd.notna(df.iloc[i, order_col]):
                            try:
                                total_sales = int(float(df.iloc[i, order_col]))
                            except:
                                pass
                        break
            
            if total_settlement > 0:
                last_day = calendar.monthrange(year, month)[1]
                data.append({
                    "date": f"{year}-{month:02d}-{last_day:02d}",
                    "amount": total_settlement,
                    "vendor_name": "배달의민족",
                    "note": f"배민 {month}월 정산 (요약)",
                    "payment_type": "delivery",
                })
        
        if not data:
            return {"status": "error", "message": "정산명세서에서 입금 데이터를 찾을 수 없습니다."}
        
        total_fees = total_sales - total_settlement if total_sales > 0 else 0
        
        return {
            "status": "success",
            "file_type": "delivery_settlement",
            "label": "🛵 배달앱 정산 (배달의민족)",
            "data": data,
            "summary": {
                "total_amount": total_settlement,
                "total_sales": total_sales,
                "total_fees": total_fees,
                "record_count": len(data),
                "channel": "배달의민족",
                "period": f"{year}년 {month}월",
            }
        }

    def _parse_yogiyo_settlement(self, df):
        """
        Parse 요기요 정산내역 (per-order format).
        Columns: 상호명, 주문번호, 주문구분, 주문일시, 주문금액, 배달료, 수수료 columns.
        Aggregates by date, calculates settlement per order.
        """
        import re, datetime
        
        # Find header rows (may be multi-row headers, rows 0-1)
        # Find key column indices
        date_col = None
        amount_col = None
        fee_cols = []     # columns to subtract (수수료, 할인 가게부담)
        credit_cols = []  # columns to add back (요기요부담할인)
        order_type_col = None
        
        # Scan first 3 rows for headers
        for i in range(min(3, len(df))):
            for j in range(len(df.columns)):
                val = df.iloc[i, j]
                if pd.isna(val):
                    continue
                cell = str(val).strip()
                
                if '주문일시' in cell:
                    date_col = j
                elif cell == '주문금액' and amount_col is None:
                    amount_col = j
                elif '주문합계' in cell and amount_col is None:
                    amount_col = j
                elif '주문구분' in cell:
                    order_type_col = j
                # Fee columns (to subtract)
                elif '이용료' in cell or ('할인' in cell and '가게 부담' in cell):
                    fee_cols.append(j)
                elif '사장님 자체할인' in cell or '사장님포인트' in cell:
                    fee_cols.append(j)
                elif '사장님배달료' in cell:
                    fee_cols.append(j)
                # Credit columns (요기요 부담 = add back)
                elif '요기요부담' in cell:
                    credit_cols.append(j)
        
        if date_col is None or amount_col is None:
            return {"status": "error", "message": "요기요 정산파일에서 주문일시/주문금액 컬럼을 찾을 수 없습니다."}
        
        # Find data start row (first row with valid date after headers)
        data_start = 0
        for i in range(min(3, len(df))):
            val = df.iloc[i, date_col]
            if pd.notna(val) and str(val).strip() != '주문일시':
                try:
                    if isinstance(val, datetime.datetime):
                        data_start = i
                        break
                    datetime.datetime.strptime(str(val).strip()[:10], "%Y-%m-%d")
                    data_start = i
                    break
                except:
                    continue
        
        # Parse orders and aggregate by date
        date_totals = {}  # date_str → {sales, fees, settlement}
        
        for i in range(data_start, len(df)):
            raw_date = df.iloc[i, date_col]
            if pd.isna(raw_date):
                continue
            
            # Skip refund rows
            if order_type_col is not None:
                otype = df.iloc[i, order_type_col]
                if pd.notna(otype) and '환불' in str(otype):
                    continue
            
            # Parse date
            try:
                if isinstance(raw_date, datetime.datetime):
                    d = raw_date.date()
                elif isinstance(raw_date, datetime.date):
                    d = raw_date
                else:
                    d = datetime.datetime.strptime(str(raw_date).strip()[:10], "%Y-%m-%d").date()
            except:
                continue
            
            # Get order amount
            raw_amount = df.iloc[i, amount_col]
            if pd.isna(raw_amount):
                continue
            try:
                order_amount = float(raw_amount)
            except:
                continue
            
            # Calculate fees
            total_fees = 0
            for fc in fee_cols:
                fv = df.iloc[i, fc] if fc < len(df.columns) else None
                if pd.notna(fv):
                    try:
                        total_fees += abs(float(fv))
                    except:
                        pass
            
            # Credits (add back)
            total_credits = 0
            for cc in credit_cols:
                cv = df.iloc[i, cc] if cc < len(df.columns) else None
                if pd.notna(cv):
                    try:
                        total_credits += abs(float(cv))
                    except:
                        pass
            
            settlement = order_amount - total_fees + total_credits
            
            date_str = d.strftime("%Y-%m-%d")
            if date_str not in date_totals:
                date_totals[date_str] = {"sales": 0, "fees": 0, "settlement": 0, "orders": 0}
            date_totals[date_str]["sales"] += order_amount
            date_totals[date_str]["fees"] += total_fees
            date_totals[date_str]["settlement"] += settlement
            date_totals[date_str]["orders"] += 1
        
        if not date_totals:
            return {"status": "error", "message": "요기요 정산파일에서 주문 데이터를 찾을 수 없습니다."}
        
        # Build entries
        data = []
        total_sales = 0
        total_fees = 0
        total_settlement = 0
        total_orders = 0
        
        for date_str in sorted(date_totals.keys()):
            dt = date_totals[date_str]
            # Use sales (gross) as revenue amount; fallback to settlement if no sales data
            revenue_amount = int(dt["sales"]) if dt["sales"] > 0 else int(dt["settlement"])
            data.append({
                "date": date_str,
                "amount": revenue_amount,
                "vendor_name": "요기요",
                "note": f"요기요 정산 ({dt['orders']}건)",
                "payment_type": "delivery",
            })
            total_sales += dt["sales"]
            total_fees += dt["fees"]
            total_settlement += dt["settlement"]
            total_orders += dt["orders"]
        
        # Detect period
        dates = sorted(date_totals.keys())
        first_date = datetime.datetime.strptime(dates[0], "%Y-%m-%d")
        
        return {
            "status": "success",
            "file_type": "delivery_settlement",
            "label": "🛵 배달앱 정산 (요기요)",
            "data": data,
            "summary": {
                "total_amount": int(total_settlement),
                "total_sales": int(total_sales),
                "total_fees": int(total_fees),
                "record_count": len(data),
                "order_count": total_orders,
                "channel": "요기요",
                "period": f"{first_date.year}년 {first_date.month}월",
            }
        }

    def _parse_coupang_eats_settlement(self, df):
        """
        Parse 쿠팡이츠 정산내역.
        Supports format with 일자, 주문금액, 정산금액 columns.
        Aggregates by date, calculates total settlement per day.
        """
        import datetime
        
        header_idx = -1
        date_col = None
        amount_col = None
        sales_col = None
        
        for i in range(min(15, len(df))):
            row_vals = [str(v).strip() for v in df.iloc[i].tolist() if pd.notna(v)]
            if '일자' in row_vals and '주문번호' in row_vals:
                header_idx = i
                break

        if header_idx == -1: return {"status": "error", "message": "쿠팡이츠 정산 내역의 시작을 찾을 수 없습니다."}

        # Search across multi-row headers
        for idx in range(len(df.columns)):
            col_text = ' '.join(str(df.iloc[r, idx]).strip() for r in range(header_idx + 1) if pd.notna(df.iloc[r, idx]))
            if '일자' in col_text and date_col is None: date_col = idx
            elif '정산금액' in col_text and amount_col is None: amount_col = idx
            elif '주문금액' in col_text and sales_col is None: sales_col = idx

        if date_col is None or amount_col is None:
            # Fallback if specific "정산금액" text wasn't in the matched column, 
            # maybe "총액" or "산정후" under "정산금액" merged cell. Look at adjacent cells in row 0.
            for idx in range(len(df.columns)):
                for r in range(header_idx + 1):
                    val = str(df.iloc[r, idx]).strip() if pd.notna(df.iloc[r, idx]) else ""
                    if '정산금액' in val: amount_col = idx
                    
            if date_col is None or amount_col is None:
                return {"status": "error", "message": f"쿠팡이츠 필수 컬럼(일자/정산금액)을 찾을 수 없습니다. (발견된 컬럼: 일자={date_col}, 정산금액={amount_col})"}

        date_totals = {}
        total_sales = 0
        total_orders = 0
        first_date = None
        
        for i in range(header_idx + 1, len(df)):
            dt = df.iloc[i, date_col]
            amount = df.iloc[i, amount_col]
            sales = df.iloc[i, sales_col] if sales_col is not None else 0
            
            if pd.isna(dt) or pd.isna(amount): continue
            
            try:
                if isinstance(dt, datetime.datetime):
                    d_str = dt.strftime("%Y-%m-%d")
                    first_date = dt if first_date is None else first_date
                else:
                    d_str = str(dt).strip()[:10]
                    first_date = datetime.datetime.strptime(d_str, "%Y-%m-%d") if first_date is None else first_date
            except: continue
            
            try: amount = int(float(str(amount).replace(',', '')))
            except: continue
            
            if pd.notna(sales):
                try: sales = int(float(str(sales).replace(',', '')))
                except: sales = 0
            else:
                sales = 0

            if amount == 0 and sales == 0: continue
            
            if d_str not in date_totals: date_totals[d_str] = {"amount": 0, "sales": 0, "orders": 0}
            date_totals[d_str]["amount"] += amount
            date_totals[d_str]["sales"] += sales
            date_totals[d_str]["orders"] += 1

        data = []
        total_amount = 0
        
        for d, dt_data in sorted(date_totals.items()):
            # Use sales (gross) as revenue amount; fallback to settlement if no sales data
            revenue_amount = dt_data["sales"] if dt_data["sales"] > 0 else dt_data["amount"]
            data.append({
                "date": d,
                "amount": revenue_amount,
                "vendor_name": "쿠팡",
                "note": f"쿠팡이츠 정산 ({dt_data['orders']}건)",
                "payment_type": "delivery",
            })
            total_amount += dt_data["amount"]
            total_sales += dt_data["sales"]
            total_orders += dt_data["orders"]
            
        period = f"{first_date.year}년 {first_date.month}월" if first_date else "알수없음"
        
        return {
            "status": "success",
            "file_type": "delivery_settlement",
            "label": "🛵 배달앱 정산 (쿠팡이츠)",
            "data": data,
            "summary": {
                "total_amount": total_amount,
                "total_sales": total_sales,
                "total_fees": total_sales - total_amount if total_sales > 0 else 0,
                "record_count": len(data),
                "order_count": total_orders,
                "channel": "쿠팡",
                "period": period,
            }
        }

    def _parse_bank_deposit_for_card_fee(self, df):
        """
        Parse 신한은행 등 은행 입금내역 to calculate total card deposits.
        Extracts 입금액/맡기신금액 and considers valid card companies.
        """
        import datetime
        import re
        
        header_idx = -1
        date_col = None
        deposit_col = None
        memo_col = None

        for i in range(min(15, len(df))):
            row_vals = [str(v).replace(' ', '') for v in df.iloc[i].tolist() if pd.notna(v)]
            has_date = any('거래일' in v or '일자' in v for v in row_vals)
            has_deposit = any('입금' in v or '맡기신' in v for v in row_vals)
            if has_date and has_deposit:
                header_idx = i
                break

        if header_idx == -1:
            return {"status": "error", "message": "은행 거래내역 헤더(거래일시/입금액)를 찾을 수 없습니다."}

        cols = [str(c).replace(' ', '') if pd.notna(c) else '' for c in df.iloc[header_idx].tolist()]
        for idx, c_str in enumerate(cols):
            if '일자' in c_str or '일시' in c_str: date_col = idx
            elif '입금' in c_str or '맡기신' in c_str: deposit_col = idx
            elif any(k in c_str for k in ['기재내용', '적요', '내용', '보내신분', '거래처', '상호명']): memo_col = idx

        if date_col is None or deposit_col is None:
            return {"status": "error", "message": "은행 거래내역에서 일자 또는 입금 컬럼을 식별하지 못했습니다."}

        total_card_deposit = 0
        first_date = None
        data = []

        for i in range(header_idx + 1, len(df)):
            dt = df.iloc[i, date_col]
            deposit = df.iloc[i, deposit_col]
            memo = str(df.iloc[i, memo_col]) if memo_col is not None and pd.notna(df.iloc[i, memo_col]) else ""

            if pd.isna(dt) or pd.isna(deposit): continue

            # Parse deposit
            try: amount = int(float(str(deposit).replace(',', '')))
            except: continue
            
            if amount <= 0: continue
            
            # Extract date
            row_date_obj = None
            try:
                if isinstance(dt, datetime.datetime): 
                    row_date_obj = dt
                else: 
                    date_str = str(dt).strip()[:10]
                    date_str = re.sub(r'[^\d-]', '-', date_str)
                    row_date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d")
                
                if first_date is None: 
                    first_date = row_date_obj
            except Exception as e:
                pass
            
            if not row_date_obj: continue

            # Default Category Guessing
            memo_no_space = memo.replace(' ', '')
            card_kws = ['카드', '비씨', '케이비', '국민', '삼성', '현대', '롯데', '하나', '농협', 'BC', '우리카드', 'KB', 'SHC', 'NH']
            pay_kws = ['제로페이', '네이버페이', '애플페이', '삼성페이', '카카오페이', '페이코', 'PAYCO', 'ZEROPAY', 'NAVERPAY', 'KAKAOPAY', 'APPLEPAY', 'SAMSUNGPAY', '간편결제']
            non_card_kws = ['서울페이', '배달', '쿠팡', '카카오', '요기요', '배민']
            
            is_card = False
            if any(k in memo_no_space for k in card_kws):
                is_card = True
            elif any(k in memo_no_space for k in pay_kws):
                is_card = True  # 페이도 카드매출과 동일하게 수수료 차감 후 입금
            elif any(memo_no_space.startswith(p) for p in ['현', '우', '국']):
                is_card = True
                
            if any(k in memo_no_space for k in non_card_kws):
                is_card = False

            default_cat = "카드수수료" if is_card else ("무시" if any(k in memo_no_space for k in non_card_kws) else "?")

            data.append({
                "date": row_date_obj.strftime("%Y-%m-%d"),
                "amount": amount,
                "memo": memo.strip(),
                "default_category": default_cat
            })

        if not data:
            return {"status": "error", "message": "유효한 입금 내역이 없습니다."}

        period = f"{first_date.year}년 {first_date.month}월"

        return {
            "status": "success",
            "file_type": "bank_deposit_card",
            "label": "🏦 은행 입금내역",
            "data": data, 
            "summary": {
                "record_count": len(data), 
                "channel": "은행입금",
                "period": period,
                "month": first_date.month,
                "year": first_date.year
            }
        }

    def _parse_ddangyo_settlement(self, df):
        """
        Parse 땡겨요 정산내역.
        Supports two formats:
          Format A (주문별): 35 columns, date split into 년도/월/일 cols 3-5
          Format B (일별): 19 columns, 입금(예정)일 single date col, 주문기간 for order dates
        """
        import re, datetime

        # Extract year and month from title row (row 0)
        year, month = None, None
        title = str(df.iloc[0, 0]) if pd.notna(df.iloc[0, 0]) else ''
        m = re.search(r'(\d{4})년\s*(\d{1,2})월', title)
        if m:
            year, month = int(m.group(1)), int(m.group(2))

        if not year or not month:
            return {"status": "error", "message": "땡겨요 정산파일에서 년/월 정보를 찾을 수 없습니다."}

        # ── Detect format and find header/data rows ──
        data_start = None
        format_type = None  # 'per_order' or 'daily'

        # Column indices (set by format detection)
        date_cols = None       # For per-order: (year_col, month_col, day_col)
        date_col = None        # For daily: single column index
        order_period_col = None # For daily: 주문기간 column
        settlement_col = None
        order_amount_col = None
        deposit_col = None     # For daily: 입금금액

        for i in range(min(len(df), 50)):
            row_text = ' '.join([str(v) for v in df.iloc[i].tolist() if pd.notna(v)])

            # Format A: per-order (2월 형식) — has 년도/월/일 columns
            if '년도' in row_text and '주문금액' in row_text:
                format_type = 'per_order'
                for j in range(len(df.columns)):
                    val = str(df.iloc[i, j]).strip() if pd.notna(df.iloc[i, j]) else ''
                    if val == '년도':
                        year_col_idx = j
                    elif val == '월' and date_cols is None:
                        month_col_idx = j
                    elif val == '일' and date_cols is None:
                        day_col_idx = j
                        date_cols = (year_col_idx, month_col_idx, day_col_idx)
                    elif val == '(C)정산금액':
                        settlement_col = j
                    elif val == '주문금액' and order_amount_col is None:
                        order_amount_col = j
                data_start = i + 1
                break

            # Format B: daily (1월 형식) — has 입금(예정)일/입금상태/주문기간
            if '입금(예정)일' in row_text and '입금상태' in row_text and '주문기간' in row_text:
                format_type = 'daily'
                for j in range(len(df.columns)):
                    val = str(df.iloc[i, j]).strip() if pd.notna(df.iloc[i, j]) else ''
                    if val == '입금(예정)일':
                        date_col = j
                    elif val == '주문기간':
                        order_period_col = j
                    elif val == '(C)정산금액':
                        settlement_col = j
                    elif val == '주문금액' and order_amount_col is None:
                        order_amount_col = j
                    elif val == '입금금액':
                        deposit_col = j
                data_start = i + 1
                break

        if data_start is None or settlement_col is None:
            return {"status": "error", "message": "땡겨요 정산파일에서 상세 데이터 헤더를 찾을 수 없습니다."}

        # ── Parse data rows ──
        date_totals = {}  # date_str → {settlement, sales, orders}

        for i in range(data_start, len(df)):
            # Skip 합 계 row
            first_cell = str(df.iloc[i, 0]).strip() if pd.notna(df.iloc[i, 0]) else ''
            if '합 계' in first_cell or '합계' in first_cell:
                break

            try:
                if format_type == 'per_order' and date_cols:
                    # Format A: separate year/month/day columns
                    y_col, m_col, d_col = date_cols
                    yr = str(df.iloc[i, y_col]).strip() if pd.notna(df.iloc[i, y_col]) else ''
                    mn = str(df.iloc[i, m_col]).strip() if pd.notna(df.iloc[i, m_col]) else ''
                    dy = str(df.iloc[i, d_col]).strip() if pd.notna(df.iloc[i, d_col]) else ''
                    if not yr or not mn or not dy:
                        continue
                    date_str = f"{int(float(yr))}-{int(float(mn)):02d}-{int(float(dy)):02d}"

                elif format_type == 'daily' and date_col is not None:
                    # Format B: 입금(예정)일 like "2026-01-16"
                    # Always use 입금(예정)일 as the revenue date (per store accounting policy)
                    raw_date = df.iloc[i, date_col]
                    if pd.isna(raw_date) or str(raw_date).strip() == '':
                        continue
                    date_str = str(raw_date).strip()[:10]
                else:
                    continue

                # Validate date
                datetime.datetime.strptime(date_str, "%Y-%m-%d")
            except Exception:
                continue

            # Get settlement amount (use 입금금액 for daily format if available, else 정산금액)
            if format_type == 'daily' and deposit_col is not None:
                settle_val = df.iloc[i, deposit_col]
            else:
                settle_val = df.iloc[i, settlement_col]
            try:
                settlement = int(float(settle_val)) if pd.notna(settle_val) else 0
            except:
                settlement = 0

            # Get order amount if available
            sales = 0
            if order_amount_col is not None:
                oa_val = df.iloc[i, order_amount_col]
                try:
                    sales = int(float(oa_val)) if pd.notna(oa_val) else 0
                except:
                    pass

            if date_str not in date_totals:
                date_totals[date_str] = {"settlement": 0, "sales": 0, "orders": 0}
            date_totals[date_str]["settlement"] += settlement
            date_totals[date_str]["sales"] += sales
            date_totals[date_str]["orders"] += 1

        if not date_totals:
            return {"status": "error", "message": "땡겨요 정산파일에서 주문 데이터를 찾을 수 없습니다."}

        # Build data entries
        data = []
        total_sales = 0
        total_settlement = 0
        total_orders = 0

        for date_str in sorted(date_totals.keys()):
            dt = date_totals[date_str]
            # Use sales (gross) as revenue amount; fallback to settlement if no sales data
            revenue_amount = dt["sales"] if dt["sales"] > 0 else dt["settlement"]
            data.append({
                "date": date_str,
                "amount": revenue_amount,
                "vendor_name": "땡겨요",
                "note": f"땡겨요 정산 ({dt['orders']}건)",
                "payment_type": "delivery",
            })
            total_sales += dt["sales"]
            total_settlement += dt["settlement"]
            total_orders += dt["orders"]

        total_fees = total_sales - total_settlement if total_sales > 0 else 0

        return {
            "status": "success",
            "file_type": "delivery_settlement",
            "label": "🛵 배달앱 정산 (땡겨요)",
            "data": data,
            "summary": {
                "total_amount": total_settlement,
                "total_sales": total_sales,
                "total_fees": total_fees,
                "record_count": len(data),
                "order_count": total_orders,
                "channel": "땡겨요",
                "period": f"{year}년 {month}월",
            }
        }


    def _parse_delivery_settlement(self, df):
        """
        Parse delivery app settlement files with packed single-column format.
        Supports: 쿠팡이츠, and similar formats.
        
        Format: "1. 2026.02.27인출-724,366원0원"
        '인출' entries = actual money deposited to bank = revenue for P&L.
        '기본정산' = Coupang-side receipt (not yet deposited), skipped.
        """
        import re
        
        # Regex: number. date type amount원 balance원
        pattern = re.compile(
            r'\d+\.\s*'
            r'(\d{4}\.\d{2}\.\d{2})'
            r'(기본정산|인출|보정|기타)'
            r'(-?[\d,]+)원'
        )
        
        daily_revenue = {}
        total_count = 0
        
        for _, row in df.iterrows():
            text = str(row.iloc[0])
            match = pattern.search(text)
            if not match:
                continue
            
            date_str = match.group(1).replace('.', '-')  # 2026.02.27 -> 2026-02-27
            tx_type = match.group(2)
            amount_str = match.group(3).replace(',', '')
            
            try:
                amount = int(amount_str)
            except ValueError:
                continue
            
            # 인출 = actual deposit to bank account = revenue for P&L
            # Amount is negative in source, so use abs()
            if tx_type == '인출':
                daily_revenue[date_str] = daily_revenue.get(date_str, 0) + abs(amount)
                total_count += 1
        
        results = []
        total_amount = 0
        date_range = [None, None]
        
        for date_str in sorted(daily_revenue.keys()):
            amount = daily_revenue[date_str]
            if date_range[0] is None:
                date_range[0] = date_str
            date_range[1] = date_str
            
            results.append({
                'date': date_str,
                'amount': amount,
                'vendor_name': '쿠팡이츠',
                'note': '쿠팡이츠 정산',
                'payment_type': 'delivery',
            })
            total_amount += amount
        
        return {
            "status": "success",
            "file_type": "delivery_settlement",
            "file_type_label": "🛵 배달앱 정산 (쿠팡이츠)",
            "data": results,
            "summary": {
                "total_amount": total_amount,
                "record_count": len(results),
                "transaction_count": total_count,
                "date_range": date_range,
            }
        }

    # ─── Keyword sets for universal column detection ───
    DATE_KEYWORDS = ['일자', '날짜', '일시', 'date', '매출일']
    AMOUNT_KEYWORDS = ['금액', '매출', '승인금액', '매출액', '결제금액', '승인액', 'amount']
    CARD_CORP_KEYWORDS = ['카드사', '매입사', '카드사명', '매입사명', '매입카드사', '발급사', '카드종류']
    CASH_KEYWORDS = ['현금', 'cash', '현금매출']
    CARD_TOTAL_KEYWORDS = ['카드매출', '카드합계', '카드', '신용카드']
    STATUS_KEYWORDS = ['구분', '승인구분', '상태', '거래구분', '유형']
    TOTAL_KEYWORDS = ['총매출', '총액', '합계', '순매출', '총합', 'total']

    def _parse_universal_revenue(self, df, file_contents):
        """
        Universal revenue file parser — works with any POS vendor format.
        
        Strategy:
        1. Auto-detect header row by scanning for date/amount keywords
        2. Map columns flexibly using broad keyword sets
        3. Classify file as 'daily_summary' or 'card_detail' by data pattern
        4. Parse and return structured data
        """
        import io
        
        # Step 1: Find header row
        header_row = -1
        for i in range(min(10, len(df))):
            row_vals = [str(v).strip() for v in df.iloc[i].tolist() if pd.notna(v)]
            row_text = ' '.join(row_vals)
            has_date = any(k in row_text for k in self.DATE_KEYWORDS)
            has_amount = any(k in row_text for k in self.AMOUNT_KEYWORDS)
            if has_date and has_amount:
                header_row = i
                break
        
        if header_row < 0:
            return {"status": "error", "message": "날짜/금액 컬럼을 자동 감지할 수 없습니다. 헤더 행에 '일자', '금액' 등의 키워드가 포함되어야 합니다."}
        
        # Re-read with proper header
        new_header = [str(v).strip() if pd.notna(v) else f'col_{j}' for j, v in enumerate(df.iloc[header_row])]
        data_df = df.iloc[header_row + 1:].copy()
        data_df.columns = new_header
        cols = list(data_df.columns)
        
        # Step 2: Map columns using keyword matching
        def find_col(keywords, exclude=None):
            for c in cols:
                cl = c.lower() if c else ''
                if any(k in c or k in cl for k in keywords):
                    if exclude and any(ex in c for ex in exclude):
                        continue
                    return c
            return None
        
        date_col = find_col(self.DATE_KEYWORDS)
        amount_col = find_col(self.AMOUNT_KEYWORDS, exclude=['현금', 'cash'])
        card_corp_col = find_col(self.CARD_CORP_KEYWORDS)
        status_col = find_col(self.STATUS_KEYWORDS)
        
        # Look for separate cash/card total columns (daily summary files)
        cash_col = find_col(self.CASH_KEYWORDS)
        card_total_col = find_col(self.CARD_TOTAL_KEYWORDS, exclude=['카드사'])
        total_col = find_col(self.TOTAL_KEYWORDS)
        
        if not date_col:
            return {"status": "error", "message": f"날짜 컬럼을 찾을 수 없습니다. 감지된 컬럼: {cols[:15]}"}
        if not amount_col and not total_col and not (cash_col or card_total_col):
            return {"status": "error", "message": f"금액 컬럼을 찾을 수 없습니다. 감지된 컬럼: {cols[:15]}"}
        
        print(f"[Universal] date={date_col}, amount={amount_col}, card_corp={card_corp_col}, "
              f"status={status_col}, cash={cash_col}, card_total={card_total_col}, total={total_col}")

        # Step 3: Classify file type by data pattern
        # Count date frequency — daily summary has ~1 row/date, detail has many
        date_counts = {}
        sample_size = min(100, len(data_df))
        for i in range(sample_size):
            dv = data_df.iloc[i].get(date_col)
            if pd.notna(dv):
                ds = str(dv)[:10]
                date_counts[ds] = date_counts.get(ds, 0) + 1
        
        avg_rows_per_date = (sum(date_counts.values()) / len(date_counts)) if date_counts else 1
        
        # If card company column exists AND avg > 2 rows per date → card transaction detail
        # If separate cash/card columns exist OR avg ≈ 1 → daily summary
        is_daily_summary = (cash_col or card_total_col or total_col) and not card_corp_col and avg_rows_per_date < 3
        is_card_detail = card_corp_col is not None and avg_rows_per_date >= 2
        
        # Fallback: if card corp column exists, treat as detail regardless
        if card_corp_col and not is_daily_summary:
            is_card_detail = True
        
        if is_daily_summary:
            return self._parse_universal_daily_summary(data_df, date_col, cash_col, card_total_col, total_col)
        elif is_card_detail:
            return self._parse_universal_card_detail(data_df, date_col, amount_col, card_corp_col, status_col)
        else:
            # Generic: treat as card detail if possible, otherwise daily
            if amount_col:
                return self._parse_universal_card_detail(data_df, date_col, amount_col, card_corp_col, status_col)
            return {"status": "error", "message": f"파일 유형을 분류할 수 없습니다. 컬럼: {cols[:15]}"}

    def _parse_universal_daily_summary(self, df, date_col, cash_col, card_col, total_col):
        """Parse a daily summary file (one row per day, cash/card breakdown)."""
        results = []
        total_amount = 0
        date_range = [None, None]
        
        def clean_num(val):
            if pd.isna(val): return 0
            try: return int(float(str(val).replace(',', '').replace('원', '')))
            except: return 0
        
        for _, row in df.iterrows():
            date_val = row.get(date_col)
            if pd.isna(date_val): continue
            
            if isinstance(date_val, datetime.datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                ds = str(date_val).strip().replace('.', '-').replace('/', '-')
                if len(ds) >= 10:
                    date_str = ds[:10]
                elif len(ds) >= 8 and ds[:4].isdigit():
                    date_str = f"{ds[:4]}-{ds[4:6]}-{ds[6:8]}"
                else:
                    continue
            
            # Validate date
            try:
                datetime.datetime.strptime(date_str[:10], '%Y-%m-%d')
            except:
                continue
            
            if date_range[0] is None: date_range[0] = date_str
            date_range[1] = date_str
            
            cash = clean_num(row.get(cash_col)) if cash_col else 0
            card = clean_num(row.get(card_col)) if card_col else 0
            total = clean_num(row.get(total_col)) if total_col else 0
            
            # If only total is available (no cash/card breakdown)
            if total > 0 and cash == 0 and card == 0:
                card = total  # Assume total is card if no breakdown
            
            # If we have total but no separate, use ratio
            if total > 0 and (cash + card) > 0 and abs(total - (cash + card)) > 100:
                # Total includes VAT or fees — scale proportionally
                ratio = total / (cash + card) if (cash + card) > 0 else 1
                cash = int(cash * ratio)
                card = int(card * ratio)
            
            if cash > 0:
                results.append({
                    'date': date_str, 'amount': cash,
                    'vendor_name': self.CASH_VENDOR_NAME,
                    'note': '현금매출', 'payment_type': 'cash',
                })
            if card > 0:
                results.append({
                    'date': date_str, 'amount': card,
                    'vendor_name': '카드매출(통합)',
                    'note': '카드매출', 'payment_type': 'card',
                })
            total_amount += (cash + card)
        
        return {
            "status": "success",
            "file_type": "pos_daily",
            "file_type_label": "📊 일자별 매출내역 (자동감지)",
            "data": results,
            "summary": {
                "total_amount": total_amount,
                "record_count": len(results),
                "date_range": date_range,
            }
        }
    
    def _parse_universal_card_detail(self, df, date_col, amount_col, card_corp_col, status_col):
        """Parse a card transaction detail file (many rows per day)."""
        daily_card = {}
        total_count = 0
        cancel_count = 0
        
        for _, row in df.iterrows():
            date_val = row.get(date_col)
            amount_val = row.get(amount_col) if amount_col else None
            
            if pd.isna(date_val) or (amount_val is not None and pd.isna(amount_val)):
                continue
            
            # Parse date
            if isinstance(date_val, datetime.datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                ds = str(date_val).strip().replace('.', '-').replace('/', '-')
                if '-' in ds:
                    date_str = ds[:10]
                elif len(ds) >= 8 and ds[:4].isdigit():
                    date_str = f"{ds[:4]}-{ds[4:6]}-{ds[6:8]}"
                else:
                    continue
            
            # Validate date
            try:
                datetime.datetime.strptime(date_str[:10], '%Y-%m-%d')
            except:
                continue
            
            # Parse amount
            try:
                amt = int(float(str(amount_val).replace(',', '').replace('원', '')))
            except (ValueError, TypeError):
                continue
            
            # Check cancellation
            if status_col:
                tx_type = str(row.get(status_col, '')).strip()
                if '취소' in tx_type:
                    amt = -amt
                    cancel_count += 1
                else:
                    total_count += 1
            else:
                total_count += 1
            
            # Card company name
            if card_corp_col and pd.notna(row.get(card_corp_col)):
                card_name = str(row.get(card_corp_col)).strip()
            else:
                card_name = '기타카드'
            
            key = (date_str, card_name)
            daily_card[key] = daily_card.get(key, 0) + amt
        
        results = []
        total_amount = 0
        date_range = [None, None]
        
        for (date_str, card_name), amount in sorted(daily_card.items()):
            if amount <= 0:
                continue
            
            vendor_name = self.CARD_VENDOR_MAP.get(card_name, f'기타카드({card_name})')
            
            if date_range[0] is None: date_range[0] = date_str
            date_range[1] = date_str
            
            results.append({
                'date': date_str, 'amount': amount,
                'vendor_name': vendor_name,
                'note': f'카드매출({card_name})',
                'payment_type': 'card',
                'card_company': card_name,
            })
            total_amount += amount
        
        return {
            "status": "success",
            "file_type": "card_detail",
            "file_type_label": "💳 카드매출 상세 (자동감지)",
            "data": results,
            "summary": {
                "total_amount": total_amount,
                "record_count": len(results),
                "transaction_count": total_count,
                "cancel_count": cancel_count,
                "date_range": date_range,
            }
        }

    def _parse_pos_card_detail_revenue(self, df):
        """Parse POS system card sales detail file (신용카드 매출내역).
        Columns: NO, 구분, 영업일자, 거래일자, 거래시간, 포스번호, 승인번호,
                 카드번호, 카드사명, 매입사명, 승인금액, 할부, 승인구분, ...
        """
        # Find header row
        header_row = 0
        for i in range(min(5, len(df))):
            row_vals = [str(v) for v in df.iloc[i].tolist() if pd.notna(v)]
            row_text = ' '.join(row_vals)
            if ('영업일자' in row_text or '거래일자' in row_text) and '승인금액' in row_text:
                header_row = i
                break

        # Set header
        new_header = [str(v).strip() if pd.notna(v) else f'col_{j}' for j, v in enumerate(df.iloc[header_row])]
        data_df = df.iloc[header_row + 1:].copy()
        data_df.columns = new_header

        # Resolve column names
        date_col = None
        for c in ['영업일자', '거래일자', '승인일자']:
            if c in data_df.columns:
                date_col = c
                break

        card_col = None
        for c in ['카드사명', '매입사명', '매입카드사']:
            if c in data_df.columns:
                card_col = c
                break

        status_col = None
        for c in ['구분', '승인구분']:
            if c in data_df.columns:
                status_col = c
                break

        if not date_col or '승인금액' not in data_df.columns:
            return {"status": "error", "message": "신용카드 매출내역 컬럼을 인식할 수 없습니다."}

        daily_card = {}
        total_count = 0
        cancel_count = 0

        for _, row in data_df.iterrows():
            date_val = row.get(date_col)
            amount_val = row.get('승인금액')

            if pd.isna(date_val) or pd.isna(amount_val):
                continue

            # Parse date
            if isinstance(date_val, datetime.datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                ds = str(date_val).strip()
                if '-' in ds:
                    date_str = ds[:10]
                elif len(ds) >= 8 and ds[:4].isdigit():
                    date_str = f"{ds[:4]}-{ds[4:6]}-{ds[6:8]}"
                else:
                    continue

            # Parse amount
            try:
                amt = int(float(str(amount_val).replace(',', '')))
            except (ValueError, TypeError):
                continue

            # Check cancellation
            tx_type = str(row.get(status_col, '')).strip() if status_col else ''
            if '취소' in tx_type:
                amt = -amt
                cancel_count += 1
            else:
                total_count += 1

            card_name = str(row.get(card_col, '기타카드')).strip() if card_col else '기타카드'
            key = (date_str, card_name)
            daily_card[key] = daily_card.get(key, 0) + amt

        results = []
        total_amount = 0
        date_range = [None, None]

        for (date_str, card_name), amount in sorted(daily_card.items()):
            if amount <= 0:
                continue

            vendor_name = self.CARD_VENDOR_MAP.get(card_name, f'기타카드({card_name})')

            if date_range[0] is None:
                date_range[0] = date_str
            date_range[1] = date_str

            results.append({
                'date': date_str,
                'amount': amount,
                'vendor_name': vendor_name,
                'note': f'카드매출({card_name})',
                'payment_type': 'card',
                'card_company': card_name,
            })
            total_amount += amount

        return {
            "status": "success",
            "file_type": "card_detail",
            "file_type_label": "💳 신용카드 매출내역 (POS)",
            "data": results,
            "summary": {
                "total_amount": total_amount,
                "record_count": len(results),
                "transaction_count": total_count,
                "cancel_count": cancel_count,
                "date_range": date_range,
            }
        }

    def _parse_pos_daily_revenue(self, df):
        """Parse POS daily revenue file (File 3 format)."""
        # Find header row with '매출일자'
        header_row = 0
        for i in range(min(5, len(df))):
            row_vals = [str(v) for v in df.iloc[i].tolist() if pd.notna(v)]
            if '매출일자' in row_vals:
                header_row = i
                break
        
        results = []
        total_amount = 0
        date_range = [None, None]
        data_start = header_row + 2  # skip main + sub header
        
        for i in range(data_start, len(df)):
            row = df.iloc[i]
            date_val = row.iloc[0]
            
            if pd.isna(date_val) or str(date_val).startswith('합'):
                continue
            
            if isinstance(date_val, datetime.datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                date_str = str(date_val)[:10]
            
            total = int(row.iloc[6]) if pd.notna(row.iloc[6]) and row.iloc[6] != 0 else 0
            cash_net = int(row.iloc[14]) if pd.notna(row.iloc[14]) and row.iloc[14] != 0 else 0
            card_net = int(row.iloc[15]) if pd.notna(row.iloc[15]) and row.iloc[15] != 0 else 0
            
            if total == 0:
                continue
            
            # FIX: User wants Gross Sales (inc. VAT).
            # iloc[6] is Total Gross. iloc[14/15] are Net.
            # We calculate ratio to scale Net -> Gross.
            net_sum = cash_net + card_net
            vat_ratio = (total / net_sum) if net_sum > 0 else 1.0
            
            # If ratio is close to 1.1, apply it. If it's 1.0, data is already Gross.
            # Just apply generally to align with Total Gross.
            cash = int(cash_net * vat_ratio)
            card = int(card_net * vat_ratio)

            if date_range[0] is None:
                date_range[0] = date_str
            date_range[1] = date_str
            
            if cash > 0:
                results.append({
                    'date': date_str,
                    'amount': cash,
                    'vendor_name': self.CASH_VENDOR_NAME,
                    'note': '현금매출',
                    'payment_type': 'cash',
                })
            
            if card > 0:
                results.append({
                    'date': date_str,
                    'amount': card,
                    'vendor_name': '카드매출(통합)',
                    'note': '카드매출',
                    'payment_type': 'card',
                })
            
            total_amount += (cash + card)
        
        return {
            "status": "success",
            "file_type": "pos_daily",
            "file_type_label": "📊 POS 일자별 매출내역",
            "data": results,
            "summary": {
                "total_amount": total_amount,
                "record_count": len(results),
                "date_range": date_range,
            }
        }

    def _parse_card_detail_revenue(self, df):
        """Parse card transaction detail file (File 2 format)."""
        header_row = 0
        for i in range(min(5, len(df))):
            row_vals = [str(v) for v in df.iloc[i].tolist() if pd.notna(v)]
            if 'No.' in row_vals or ('구분' in row_vals and '카드사' in row_vals):
                header_row = i
                break
        
        data_start = header_row + 1
        daily_card = {}
        total_count = 0
        cancel_count = 0
        
        for i in range(data_start, len(df)):
            row = df.iloc[i]
            date_val = row.iloc[2]
            card_company = row.iloc[4]
            tx_type = row.iloc[1]
            amount = row.iloc[8]
            
            if pd.isna(date_val) or pd.isna(amount):
                continue
            
            if isinstance(date_val, datetime.datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                date_str = str(date_val)[:10]
            
            try:
                amt = int(float(str(amount).replace(',', '')))
            except (ValueError, TypeError):
                continue
            
            if str(tx_type) == '취소':
                amt = -amt
                cancel_count += 1
            else:
                total_count += 1
            
            card_name = str(card_company).strip() if pd.notna(card_company) else '기타카드'
            key = (date_str, card_name)
            daily_card[key] = daily_card.get(key, 0) + amt
        
        results = []
        total_amount = 0
        date_range = [None, None]
        
        for (date_str, card_name), amount in sorted(daily_card.items()):
            if amount <= 0:
                continue
            
            vendor_name = self.CARD_VENDOR_MAP.get(card_name, f'기타카드({card_name})')
            
            if date_range[0] is None:
                date_range[0] = date_str
            date_range[1] = date_str
            
            results.append({
                'date': date_str,
                'amount': amount,
                'vendor_name': vendor_name,
                'note': f'카드매출({card_name})',
                'payment_type': 'card',
                'card_company': card_name,
            })
            total_amount += amount
        
        return {
            "status": "success",
            "file_type": "card_detail",
            "file_type_label": "💳 카드 상세매출내역",
            "data": results,
            "summary": {
                "total_amount": total_amount,
                "record_count": len(results),
                "transaction_count": total_count,
                "cancel_count": cancel_count,
                "date_range": date_range,
            }
        }

    def _parse_card_summary_revenue(self, df):
        """Parse monthly card summary file (File 1 format)."""
        results = []
        total_amount = 0
        
        for i in range(len(df)):
            row = df.iloc[i]
            val = str(row.iloc[0])
            
            if len(val) >= 7 and val[4] == '-' and val[:4].isdigit():
                year_month = val[:7]
                amount = row.iloc[1]
                
                if pd.notna(amount):
                    try:
                        amt = int(float(str(amount).replace(',', '')))
                    except (ValueError, TypeError):
                        continue
                    
                    approval_total = int(float(str(row.iloc[3]).replace(',', ''))) if pd.notna(row.iloc[3]) else 0
                    approval_count = int(float(str(row.iloc[4]).replace(',', ''))) if pd.notna(row.iloc[4]) else 0
                    cancel_total = int(float(str(row.iloc[5]).replace(',', ''))) if pd.notna(row.iloc[5]) else 0
                    cancel_count = int(float(str(row.iloc[6]).replace(',', ''))) if pd.notna(row.iloc[6]) else 0
                    
                    results.append({
                        'year_month': year_month,
                        'net_amount': amt,
                        'approval_total': approval_total,
                        'approval_count': approval_count,
                        'cancel_total': cancel_total,
                        'cancel_count': cancel_count,
                    })
                    total_amount += amt
        
        return {
            "status": "success",
            "file_type": "card_summary",
            "file_type_label": "📈 월별 카드매출 요약",
            "data": results,
            "summary": {
                "total_amount": total_amount,
                "months": len(results),
            },
            "message": "월별 카드매출 요약 데이터입니다. 상세 매출 데이터를 입력하려면 '카드상세매출내역' 또는 'POS 일자별 매출' 파일을 업로드해 주세요."
        }
