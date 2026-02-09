"""
배달앱 매출 파서 서비스
쿠팡이츠, 요기요, 땡겨요, 배민 정산 엑셀 파일 파싱
"""
import pandas as pd
import io
import re
import json
import datetime
from typing import Dict, Any, List, Optional


def safe_int(val) -> int:
    """Safely convert value to int."""
    if val is None or val == '' or (isinstance(val, float) and pd.isna(val)):
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def extract_year_month_from_filename(filename: str) -> tuple:
    """Extract year and month from filename patterns."""
    # Pattern: 2026-01, 2026년 1월, 202601, 2026_01
    patterns = [
        r'(\d{4})-(\d{1,2})',
        r'(\d{4})년\s*(\d{1,2})월',
        r'(\d{4})(\d{2})_',
        r'(\d{4})_(\d{2})',
        r'(\d{4})년_(\d{2})월',
    ]
    for pat in patterns:
        m = re.search(pat, filename)
        if m:
            return int(m.group(1)), int(m.group(2))
    return None, None


# ═══════════════════════════════════════════════════════════════
#  쿠팡이츠 Parser
# ═══════════════════════════════════════════════════════════════

def parse_coupang(filepath_or_bytes, filename: str = "") -> Dict[str, Any]:
    """
    쿠팡이츠 정산 엑셀 파싱.
    Row 0-2: multi-row header
    Row 3+: individual order data
    Key columns (0-indexed):
      0: 일자, 9: 총금액, 10: 주문금액,
      14-16: 중개이용료, 17-18: 결제대행수수료, 19-21: 배달비,
      26-33: 서비스이용료, 34-36: 광고비, 37-39: 정산금액
    """
    if isinstance(filepath_or_bytes, bytes):
        df = pd.read_excel(io.BytesIO(filepath_or_bytes), header=None)
    else:
        df = pd.read_excel(filepath_or_bytes, header=None)

    year, month = extract_year_month_from_filename(filename)

    # Data starts at row 3
    data_rows = df.iloc[3:]

    total_sales = 0      # 총금액 (col 9)
    order_amount = 0     # 주문금액 (col 10)
    total_settlement = 0 # 정산금액 (col 39 = 산정후)
    order_count = 0
    fee_mediation = 0    # 중개이용료 (col 16)
    fee_payment = 0      # 결제대행수수료 (col 18 if exists, else col 17)
    fee_delivery = 0     # 배달비 (col 21)
    fee_service = 0      # 서비스이용료 (col 30 or 33)
    fee_ad = 0           # 광고비 (col 36)
    fee_promo_shop = 0   # 상점부담쿠폰 (col 13)

    for _, row in data_rows.iterrows():
        # Skip cancelled or non-data rows
        tx_type = str(row.iloc[8]) if len(row) > 8 else ''
        date_val = row.iloc[0]
        if pd.isna(date_val) or str(date_val).strip() == '':
            continue

        # Detect year/month from first data row if not in filename
        if year is None and date_val:
            try:
                dt = pd.to_datetime(date_val)
                year, month = dt.year, dt.month
            except:
                pass

        total_sales += safe_int(row.iloc[9]) if len(row) > 9 else 0
        order_amount += safe_int(row.iloc[10]) if len(row) > 10 else 0

        # 정산금액: col 39 (산정후) or col 37 (산정전)
        if len(row) > 39:
            total_settlement += safe_int(row.iloc[39])
        elif len(row) > 37:
            total_settlement += safe_int(row.iloc[37])

        # Fee breakdowns
        fee_mediation += safe_int(row.iloc[16]) if len(row) > 16 else 0
        fee_payment += safe_int(row.iloc[17]) if len(row) > 17 else 0
        fee_delivery += safe_int(row.iloc[21]) if len(row) > 21 else 0
        fee_promo_shop += safe_int(row.iloc[13]) if len(row) > 13 else 0

        # 서비스이용료: col 30 (산정후)
        if len(row) > 30:
            fee_service += safe_int(row.iloc[30])

        # 광고비: col 36 (총액)
        if len(row) > 36:
            fee_ad += safe_int(row.iloc[36])

        if tx_type == '결제':
            order_count += 1
        elif tx_type == '취소':
            order_count -= 1
        else:
            order_count += 1

    total_fees = total_sales - total_settlement

    fee_breakdown = {
        "중개이용료": abs(fee_mediation),
        "결제대행수수료": abs(fee_payment),
        "배달비": abs(fee_delivery),
        "서비스이용료": abs(fee_service),
        "광고비": abs(fee_ad),
        "상점부담쿠폰": abs(fee_promo_shop),
    }

    return {
        "channel": "쿠팡",
        "year": year,
        "month": month,
        "total_sales": total_sales,
        "total_fees": abs(total_fees),
        "settlement_amount": total_settlement,
        "order_count": max(0, order_count),
        "fee_breakdown": fee_breakdown,
    }


# ═══════════════════════════════════════════════════════════════
#  요기요 Parser
# ═══════════════════════════════════════════════════════════════

def parse_yogiyo(filepath_or_bytes, filename: str = "") -> Dict[str, Any]:
    """
    요기요 정산 엑셀 파싱.
    '요약' sheet:
      Row 2: A, 주문금액, value
      Row 4: C, 차감금액, value
      Row 19: D, 정산금액, value
    '상세 거래내역' sheet: individual orders
    """
    if isinstance(filepath_or_bytes, bytes):
        xls = pd.ExcelFile(io.BytesIO(filepath_or_bytes))
    else:
        xls = pd.ExcelFile(filepath_or_bytes)

    year, month = extract_year_month_from_filename(filename)

    # Parse summary sheet
    summary_df = pd.read_excel(xls, sheet_name='요약', header=None)

    total_sales = 0
    total_fees = 0
    settlement = 0
    fee_breakdown = {}

    for _, row in summary_df.iterrows():
        label = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
        name = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ''
        value = safe_int(row.iloc[3]) if len(row) > 3 else 0

        if label == 'A' and '주문금액' in name:
            total_sales = value
        elif label == 'C' and '차감금액' in name:
            total_fees = value
        elif label == 'D' and '정산금액' in name:
            settlement = value
        # Fee items
        elif label.startswith('C-'):
            if value != 0:
                fee_breakdown[name] = abs(value)

    # Count orders from detail sheet
    order_count = 0
    try:
        detail_df = pd.read_excel(xls, sheet_name='상세 거래내역', header=None)
        # Data starts at row 3, ends before summary row
        for i in range(3, len(detail_df)):
            no_val = detail_df.iloc[i, 0]
            if pd.isna(no_val) or str(no_val).strip() == '':
                break
            order_count += 1
    except:
        pass

    return {
        "channel": "요기요",
        "year": year,
        "month": month,
        "total_sales": total_sales,
        "total_fees": abs(total_fees),
        "settlement_amount": settlement,
        "order_count": order_count,
        "fee_breakdown": fee_breakdown,
    }


# ═══════════════════════════════════════════════════════════════
#  땡겨요 Parser
# ═══════════════════════════════════════════════════════════════

def parse_ddangyo(filepath_or_bytes, filename: str = "") -> Dict[str, Any]:
    """
    땡겨요 정산 엑셀 파싱.
    Summary in rows 3-5:
      Row 4: (A)주문결제, (B)차감금액, (C)정산금액, ...
      Row 5: actual values
    Detail in rows 44+:
      Header row 44: 주문중개이용료, 결제정산이용료, 땡배달이용료, 사장님쿠폰, ...
      Last row: 합 계
    """
    if isinstance(filepath_or_bytes, bytes):
        df = pd.read_excel(io.BytesIO(filepath_or_bytes), header=None)
    else:
        df = pd.read_excel(filepath_or_bytes, header=None)

    year, month = extract_year_month_from_filename(filename)

    # Extract from title row (row 0): "2025년09월  정산 내역"
    if year is None:
        title = str(df.iloc[0, 0]) if pd.notna(df.iloc[0, 0]) else ''
        m = re.search(r'(\d{4})년(\d{1,2})월', title)
        if m:
            year, month = int(m.group(1)), int(m.group(2))

    # Summary row (row index 5)
    total_sales = safe_int(df.iloc[5, 0])       # (A)주문결제
    total_fees_neg = safe_int(df.iloc[5, 1])     # (B)차감금액 (negative)
    settlement = safe_int(df.iloc[5, 2])         # (C)정산금액

    total_fees = abs(total_fees_neg)

    # Count detail rows and extract fee breakdown from 합계 row
    order_count = 0
    fee_breakdown = {}

    # Find the data section (after headers around row 42-44)
    data_start = None
    for i in range(40, min(50, len(df))):
        cell = str(df.iloc[i, 0]) if i < len(df) and pd.notna(df.iloc[i, 0]) else ''
        if '입금' in cell and '예정' in cell:
            data_start = i + 1
            break

    if data_start:
        for i in range(data_start, len(df)):
            first_cell = str(df.iloc[i, 0]) if pd.notna(df.iloc[i, 0]) else ''
            if first_cell == '합 계':
                # Extract fee breakdown from total row
                fee_breakdown = {
                    "주문중개이용료": abs(safe_int(df.iloc[i, 10])),
                    "결제정산이용료": abs(safe_int(df.iloc[i, 11])),
                    "땡배달이용료": abs(safe_int(df.iloc[i, 12])),
                    "사장님쿠폰": abs(safe_int(df.iloc[i, 13])),
                    "프로모션사장님부담금": abs(safe_int(df.iloc[i, 14])),
                }
                break
            elif first_cell and first_cell != '':
                order_count += 1

    return {
        "channel": "땡겨요",
        "year": year,
        "month": month,
        "total_sales": total_sales,
        "total_fees": total_fees,
        "settlement_amount": settlement,
        "order_count": order_count,
        "fee_breakdown": fee_breakdown,
    }


# ═══════════════════════════════════════════════════════════════
#  배민 Parser
# ═══════════════════════════════════════════════════════════════

def parse_baemin(filepath_or_bytes, filename: str = "", password: str = "630730") -> Dict[str, Any]:
    """
    배달의민족 정산명세서 파싱.
    Encrypted OLE2 xlsx file.
    '요약' sheet: Row 5=headers, Row 6=values
      Col 1=(A)주문중개, Col 2=(B)배달, Col 3=(C)그외,
      Col 4=(D)기타, Col 5=(E)부가세, Col 6=(F)우리가게클릭,
      Col 7=(G)배민오더, Col 8=(H)입금금액
    '상세' sheet: individual order rows for order count
    """
    import msoffcrypto
    import openpyxl

    year, month = extract_year_month_from_filename(filename)

    # Decrypt
    if isinstance(filepath_or_bytes, bytes):
        f = io.BytesIO(filepath_or_bytes)
    else:
        f = open(filepath_or_bytes, 'rb')

    try:
        ms = msoffcrypto.OfficeFile(f)
        if ms.is_encrypted():
            ms.load_key(password=password)
            decrypted = io.BytesIO()
            ms.decrypt(decrypted)
            decrypted.seek(0)
            wb = openpyxl.load_workbook(decrypted, data_only=True)
        else:
            if isinstance(filepath_or_bytes, bytes):
                wb = openpyxl.load_workbook(io.BytesIO(filepath_or_bytes), data_only=True)
            else:
                wb = openpyxl.load_workbook(filepath_or_bytes, data_only=True)
    finally:
        if not isinstance(filepath_or_bytes, bytes):
            f.close()

    total_sales = 0
    settlement = 0
    order_count = 0
    fee_breakdown = {}

    # ── Parse 요약 sheet ──
    summary_ws = None
    for sn in wb.sheetnames:
        if '요약' in sn:
            summary_ws = wb[sn]
            break

    if summary_ws:
        # Find the header row with (A), (B), etc.
        header_row = None
        for r in range(1, min(15, summary_ws.max_row + 1)):
            val = summary_ws.cell(r, 1).value
            if val and '(A)' in str(val):
                header_row = r
                break

        if header_row:
            data_row = header_row + 1
            # Read headers and values
            headers = {}
            for c in range(1, summary_ws.max_column + 1):
                h = summary_ws.cell(header_row, c).value
                if h:
                    headers[c] = str(h).strip()

            for c, label in headers.items():
                val = safe_int(summary_ws.cell(data_row, c).value)
                if '주문중개' in label or '(A)' in label:
                    total_sales = val
                elif '입금금액' in label or '(H)' in label:
                    settlement = val
                elif val != 0:
                    # Fee items (B through G are all fees/adjustments)
                    clean_label = label.replace('(B)', '').replace('(C)', '').replace('(D)', '').replace('(E)', '').replace('(F)', '').replace('(G)', '').strip()
                    fee_breakdown[clean_label] = abs(val)

    # ── Count orders from 상세 sheet ──
    detail_ws = None
    for sn in wb.sheetnames:
        if '상세' in sn:
            detail_ws = wb[sn]
            break

    if detail_ws:
        # Find data start (after header rows, look for date pattern)
        for r in range(1, detail_ws.max_row + 1):
            val = detail_ws.cell(r, 1).value
            if val:
                val_str = str(val).strip()
                # Date pattern like 2026-01-02
                if re.match(r'\d{4}-\d{2}-\d{2}', val_str):
                    order_count += 1

    total_fees = abs(total_sales - settlement)

    return {
        "channel": "배민",
        "year": year,
        "month": month,
        "total_sales": total_sales,
        "total_fees": total_fees,
        "settlement_amount": settlement,
        "order_count": order_count,
        "fee_breakdown": fee_breakdown,
    }


# ═══════════════════════════════════════════════════════════════
#  Auto-detect and parse
# ═══════════════════════════════════════════════════════════════

def detect_and_parse(filepath: str, filename: str = "") -> Optional[Dict[str, Any]]:
    """Auto-detect delivery platform from filename/path and parse."""
    if not filename:
        import os
        filename = os.path.basename(filepath)

    fname_lower = filename.lower()
    parent_dir = ""
    if filepath:
        import os
        parent_dir = os.path.basename(os.path.dirname(filepath)).lower()

    if 'coupang' in fname_lower or '쿠팡' in parent_dir:
        return parse_coupang(filepath, filename)
    elif '요기요' in fname_lower or '요기요' in parent_dir:
        return parse_yogiyo(filepath, filename)
    elif '땡겨요' in fname_lower or '땡겨요' in parent_dir:
        return parse_ddangyo(filepath, filename)
    elif '배달의민족' in fname_lower or '배민' in parent_dir:
        return parse_baemin(filepath, filename)
    else:
        return None
